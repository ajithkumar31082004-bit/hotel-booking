import sys, io

# Fix Windows CP1252 UnicodeEncodeError for ₹ ✓ emoji etc.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    flash,
    jsonify,
    send_file,
    make_response,
)
from flask_babel import Babel, gettext
from decimal import Decimal

# i18n
LANGUAGES = {"en": "English", "hi": "Hindi"}
import boto3
from datetime import datetime, timedelta
import uuid
import random
from functools import wraps
import os

import hashlib
from werkzeug.utils import secure_filename
from pathlib import Path
from io import StringIO, BytesIO
import csv
import json

# ── AI / ML Engine imports ──────────────────────────────────────────────────
try:
    from ai_recommender import get_recommendations, MODEL_METADATA as REC_META

    AI_RECOMMENDER_OK = True
except ImportError:
    AI_RECOMMENDER_OK = False

try:
    from ai_demand_forecast import forecast_demand

    AI_FORECAST_OK = True
except ImportError:
    AI_FORECAST_OK = False

try:
    from ai_dynamic_pricing import apply_dynamic_pricing, compute_dynamic_price

    AI_PRICING_OK = True
except ImportError:
    AI_PRICING_OK = False

try:
    from ai_sentiment import analyze_all_reviews, analyze_review

    AI_SENTIMENT_OK = True
except ImportError:
    AI_SENTIMENT_OK = False

try:
    from ai_fraud_detection import (
        score_booking as fraud_score_booking,
        scan_all_bookings,
    )

    AI_FRAUD_OK = True
except ImportError:
    AI_FRAUD_OK = False

try:
    from ai_cancellation import (
        predict_cancellation,
        predict_all as predict_all_cancellations,
    )

    AI_CANCEL_OK = True
except ImportError:
    AI_CANCEL_OK = False

# Excel Generation imports
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel generation will use CSV fallback.")

# PDF Generation imports
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        PageBreak,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab not installed. PDF generation will be skipped.")


app = Flask(__name__)
app.secret_key = "cHOgZyQn5hyqDrlIRNqxdDvecq4Sc7sD5fgrc3aB"

# Configure upload folder
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "static", "uploads", "rooms")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

# Create upload folder if it doesn't exist
Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE
app.config["BABEL_TRANSLATION_DIRECTORIES"] = "translations"
app.config["BABEL_DEFAULT_LOCALE"] = "en"


def get_locale():
    if "language" in session:
        return session["language"]


try:
    babel = Babel(app, locale_selector=get_locale)
except TypeError:
    babel = Babel(app)
    try:
        babel.localeselector(get_locale)
    except Exception:
        pass


# Make LANGUAGES available to all templates
@app.context_processor
def inject_languages():
    return dict(LANGUAGES=LANGUAGES)


@app.template_filter("getServiceIcon")
def get_service_icon(service_type):
    """Get Font Awesome icon for service type"""
    icons = {
        "room_service": "utensils",
        "spa": "spa",
        "restaurant": "hamburger",
        "laundry": "tshirt",
        "transportation": "taxi",
        "tour": "map-marked-alt",
        "other": "concierge-bell",
    }
    return icons.get(service_type, "concierge-bell")


def convert_decimal_to_float(value):
    """Convert Decimal to float, handling None and already-float values"""
    if value is None:
        return 0.0
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    return 0.0


# Import database and notification modules
try:
    from db import (
        create_tables,
        add_user,
        get_user,
        add_room,
        get_rooms,
        add_booking,
        get_user_bookings,
        get_all_bookings,
        update_room_availability,
        get_all_rooms,
        delete_user,
        delete_room,
        get_analytics,
        get_all_users,
        add_review,
        get_room_reviews,
        get_user_reviews,
        get_all_reviews,
        get_room_rating,
        add_branch,
        get_branch,
        get_all_branches,
        get_branch_by_city,
        get_loyalty_points,
        add_loyalty_points,
        redeem_loyalty_points,
        init_indian_branches,
        get_booking,
        update_booking_status,
        delete_booking,
        update_user,
        update_room,
        add_favorite_room,
        remove_favorite_room,
        get_user_favorites,
        is_room_favorited,
        get_past_bookings,
        get_upcoming_bookings,
        calculate_total_spent,
        get_booking_stats,
        add_chat_message,
        get_user_chat_messages,
        mark_messages_as_read,
        add_chat_request,
        get_chat_requests,
        update_chat_request,
        get_chat_request,
        get_unread_messages_count,
        get_all_chat_messages,
        delete_chat_message,
        delete_chat_request,
        calculate_dynamic_price,
        get_pricing_rules,
        init_default_pricing_rules,
        save_guest_preferences,
        get_guest_preferences,
        add_to_waitlist,
        get_user_waitlist,
        get_room_waitlist,
        update_waitlist_status,
        notify_waitlist,
        get_services,
        book_service,
        get_user_service_bookings,
        init_default_services,
        add_notification,
        get_user_notifications,
        schedule_booking_reminders,
    )
    from sns_notifier import (
        subscribe_email,
        send_booking_confirmation,
        send_notification,
    )

    DB_AVAILABLE = True
except ImportError as e:
    print(f"Warning: Import error - {e}")
    # Create mock functions for testing
    DB_AVAILABLE = False

    def create_tables():
        return True

    def add_user(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def get_user(*args, **kwargs):
        return None

    def add_room(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def get_rooms(*args, **kwargs):
        return []

    def add_booking(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def get_user_bookings(*args, **kwargs):
        return []

    def get_all_bookings(*args, **kwargs):
        return []

    def update_room_availability(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def get_all_rooms(*args, **kwargs):
        return []

    def delete_user(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def delete_room(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def get_analytics(*args, **kwargs):
        return {
            "users": {"total": 0, "by_role": {}},
            "rooms": {"total": 0, "by_availability": {}},
            "bookings": {"total": 0, "by_status": {}},
        }

    def subscribe_email(*args, **kwargs):
        return "mock-subscription"

    def send_booking_confirmation(*args, **kwargs):
        return "mock-message-id"

    def get_booking(*args, **kwargs):
        return None

    def update_booking_status(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def delete_booking(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def update_user(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}

    def update_room(*args, **kwargs):
        return {"ResponseMetadata": {"HTTPStatusCode": 200}}


# Initialize database
def init_db():
    """Initialize database and populate sample data"""
    try:
        create_tables()

        # Initialize default users if DB is available
        if DB_AVAILABLE:
            init_default_users()

        # Initialize Indian branches
        if DB_AVAILABLE:
            init_indian_branches()

        # Always ensure sample rooms are created (mock or real DB)
        init_sample_rooms()

        print("Database tables initialized/verified")
        print(f"Total rooms in system: {len(get_all_rooms())}")
        print(f"Total branches: {len(get_all_branches())}")

    except Exception as e:
        print(f"Database initialization error: {e}")


def init_default_users():
    """Initialize default users for all 5 roles — Blissful Abodes Chennai"""
    try:
        default_users = [
            # Level 5: Super Admin (Owner/CEO)
            {
                "user_id": str(uuid.uuid4()),
                "name": "Vikram Patel",
                "email": "superadmin@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 45,
                "role": "super_admin",
                "phone": "+91 98765 00001",
                "created_at": datetime.now().isoformat(),
            },
            # Level 4: Admin (General Manager)
            {
                "user_id": str(uuid.uuid4()),
                "name": "Anil Sharma",
                "email": "admin@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 42,
                "role": "admin",
                "phone": "+91 98765 00002",
                "created_at": datetime.now().isoformat(),
            },
            # Level 3: Manager (Department Heads)
            {
                "user_id": str(uuid.uuid4()),
                "name": "Suresh Menon",
                "email": "manager@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 38,
                "role": "manager",
                "department": "Front Office",
                "phone": "+91 98765 00003",
                "created_at": datetime.now().isoformat(),
            },
            {
                "user_id": str(uuid.uuid4()),
                "name": "Meena Krishnan",
                "email": "hk.manager@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 36,
                "role": "manager",
                "department": "Housekeeping",
                "phone": "+91 98765 00004",
                "created_at": datetime.now().isoformat(),
            },
            # Level 2: Staff (Front Desk / Housekeeping)
            {
                "user_id": str(uuid.uuid4()),
                "name": "Priya Sharma",
                "email": "staff@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 28,
                "role": "staff",
                "department": "Front Desk",
                "shift": "Morning",
                "phone": "+91 98765 00005",
                "created_at": datetime.now().isoformat(),
            },
            {
                "user_id": str(uuid.uuid4()),
                "name": "Raj Kumar",
                "email": "staff2@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 26,
                "role": "staff",
                "department": "Housekeeping",
                "shift": "Evening",
                "phone": "+91 98765 00006",
                "created_at": datetime.now().isoformat(),
            },
            # Level 1: Guest (Customers)
            {
                "user_id": str(uuid.uuid4()),
                "name": "Rajesh Kumar",
                "email": "guest@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 35,
                "role": "guest",
                "phone": "+91 98765 43210",
                "created_at": datetime.now().isoformat(),
            },
            {
                "user_id": str(uuid.uuid4()),
                "name": "Anita Desai",
                "email": "guest2@blissfulabodes.com",
                "password": hash_password("password123"),
                "age": 30,
                "role": "guest",
                "phone": "+91 98765 43211",
                "created_at": datetime.now().isoformat(),
            },
            # Legacy accounts for backward compatibility
            {
                "user_id": str(uuid.uuid4()),
                "name": "Admin User",
                "email": "admin@example.com",
                "password": hash_password("password123"),
                "age": 35,
                "role": "admin",
                "created_at": datetime.now().isoformat(),
            },
            {
                "user_id": str(uuid.uuid4()),
                "name": "Guest User",
                "email": "guest@example.com",
                "password": hash_password("password123"),
                "age": 28,
                "role": "guest",
                "created_at": datetime.now().isoformat(),
            },
        ]

        for user in default_users:
            existing_user = get_user(user["email"])
            if not existing_user:
                add_user(user)
                print(f"Created default {user['role']}: {user['email']}")
    except Exception as e:
        print(f"Error initializing default users: {e}")


def init_sample_rooms():
    """Initialize 100 rooms for Blissful Abodes Chennai"""
    try:
        existing_rooms = get_all_rooms()
        print(f"Current rooms in database: {len(existing_rooms)}")

        if len(existing_rooms) < 100:
            print("Initializing 100 Chennai rooms...")
            create_inline_sample_rooms()
        else:
            print(f"✓ Database already contains {len(existing_rooms)} rooms")

    except Exception as e:
        print(f"Error initializing sample rooms: {e}")
        import traceback

        traceback.print_exc()


def create_inline_sample_rooms():
    """Create 100 rooms for Blissful Abodes Chennai — floors 1-7, numbered S100-V199"""
    try:
        existing_rooms = get_all_rooms()
        if len(existing_rooms) >= 100:
            print(f"Rooms already exist ({len(existing_rooms)}), skipping population")
            return

        print("Creating 100 Chennai rooms...")
        LOCATION = "Chennai, Tamil Nadu"
        room_count = 0

        # Floor 1-3: Single Rooms (S100-S124) — 25 rooms
        single_amenities = [
            "WiFi",
            "32-inch TV",
            "AC",
            "Work Desk",
            "Private Bathroom",
            "Tea/Coffee Maker",
        ]
        for i in range(25):
            room_num = 100 + i
            floor = 1 + (i // 9)
            price = random.randint(2000, 5000)
            room = {
                "room_id": str(uuid.uuid4()),
                "room_number": f"S{room_num}",
                "name": f"Single Room S{room_num}",
                "location": LOCATION,
                "price": float(price),
                "capacity": 1,
                "floor": floor,
                "amenities": single_amenities,
                "availability": "available" if random.random() > 0.3 else "occupied",
                "image": get_room_image_url("single", i),
                "room_type": "single",
                "virtual_tour_360_url": get_360_tour_image("single", i),
                "created_at": datetime.now().isoformat(),
            }
            add_room(room)
            room_count += 1

        # Floor 2-4: Double Rooms (D125-D149) — 25 rooms
        double_amenities = [
            "King/Queen Bed",
            "WiFi",
            "42-inch Smart TV",
            "AC",
            "Mini Fridge",
            "Private Bathroom",
            "Coffee Maker",
            "Seating Area",
        ]
        for i in range(25):
            room_num = 125 + i
            floor = 2 + (i // 9)
            price = random.randint(5000, 10000)
            room = {
                "room_id": str(uuid.uuid4()),
                "room_number": f"D{room_num}",
                "name": f"Deluxe Double Room D{room_num}",
                "location": LOCATION,
                "price": float(price),
                "capacity": 2,
                "floor": floor,
                "amenities": double_amenities,
                "availability": "available" if random.random() > 0.3 else "occupied",
                "image": get_room_image_url("double", i),
                "room_type": "double",
                "virtual_tour_360_url": get_360_tour_image("double", i),
                "created_at": datetime.now().isoformat(),
            }
            add_room(room)
            room_count += 1

        # Floor 4-5: Family Suites (F150-F174) — 25 rooms
        family_amenities = [
            "2 Queen Beds",
            "WiFi",
            "50-inch Smart TV",
            "AC",
            "Kitchenette",
            "Two Bathrooms",
            "Dining Table",
            "Sofa Set",
        ]
        for i in range(25):
            room_num = 150 + i
            floor = 4 + (i // 13)
            price = random.randint(10000, 20000)
            room = {
                "room_id": str(uuid.uuid4()),
                "room_number": f"F{room_num}",
                "name": f"Family Suite F{room_num}",
                "location": LOCATION,
                "price": float(price),
                "capacity": 4,
                "floor": floor,
                "amenities": family_amenities,
                "availability": "available" if random.random() > 0.3 else "occupied",
                "image": get_room_image_url("family", i),
                "room_type": "family",
                "virtual_tour_360_url": get_360_tour_image("family", i),
                "created_at": datetime.now().isoformat(),
            }
            add_room(room)
            room_count += 1

        # Floor 5-6: Couple/Romantic Rooms (C175-C189) — 15 rooms
        couple_amenities = [
            "King Bed",
            "WiFi",
            "55-inch OLED TV",
            "AC",
            "Jacuzzi",
            "Romantic Decor",
            "Private Balcony",
            "Premium Toiletries",
            "Champagne Welcome",
        ]
        romantic_names = [
            "Honeymoon Suite",
            "Romance Retreat",
            "Ocean Romance",
            "Garden Suite",
            "Marina Suite",
        ]
        for i in range(15):
            room_num = 175 + i
            floor = 5 + (i // 8)
            price = random.randint(12000, 25000)
            room = {
                "room_id": str(uuid.uuid4()),
                "room_number": f"C{room_num}",
                "name": f"{romantic_names[i % len(romantic_names)]} C{room_num}",
                "location": LOCATION,
                "price": float(price),
                "capacity": 2,
                "floor": floor,
                "amenities": couple_amenities,
                "availability": "available" if random.random() > 0.25 else "occupied",
                "image": get_room_image_url("couple", i),
                "room_type": "couple",
                "virtual_tour_360_url": get_360_tour_image("couple", i),
                "created_at": datetime.now().isoformat(),
            }
            add_room(room)
            room_count += 1

        # Floor 6-7: VIP Suites (V190-V199) — 10 rooms
        vip_amenities = [
            "Emperor King Bed",
            "WiFi",
            "65-inch OLED TV",
            "AC",
            "Private Jacuzzi",
            "Butler Service",
            "Sea View Balcony",
            "Premium Bar",
            "Complimentary Breakfast",
            "Limousine Pickup",
        ]
        vip_names = [
            "Presidential Suite",
            "Royal Suite",
            "Penthouse",
            "Executive Suite",
            "Maharaja Suite",
        ]
        for i in range(10):
            room_num = 190 + i
            floor = 6 + (i // 5)
            price = random.randint(18000, 35000)
            room = {
                "room_id": str(uuid.uuid4()),
                "room_number": f"V{room_num}",
                "name": f"{vip_names[i % len(vip_names)]} V{room_num}",
                "location": LOCATION,
                "price": float(price),
                "capacity": 2,
                "floor": floor,
                "amenities": vip_amenities,
                "availability": "available" if random.random() > 0.2 else "occupied",
                "image": get_room_image_url("vip", i),
                "room_type": "vip",
                "virtual_tour_360_url": get_360_tour_image("vip", i),
                "created_at": datetime.now().isoformat(),
            }
            add_room(room)
            room_count += 1

        print(
            f"[OK] Successfully created {room_count} rooms for Blissful Abodes Chennai"
        )

    except Exception as e:
        print(f"Error creating inline sample rooms: {e}")
        import traceback

        traceback.print_exc()


def get_room_image_url(room_type=None, room_index=0):
    """
    Get a unique room image URL from open-source image providers (Unsplash/Pexels).
    Returns different images for different room types to show variety.
    """
    # Collection of open-source room images from Unsplash (free to use)
    # Format: Unsplash Source API URLs - these are free and open-source

    room_image_collection = {
        "single": [
            "https://images.unsplash.com/photo-1631049307264-da0ec9d70304?w=800&h=600&fit=crop",  # Modern single room
            "https://images.unsplash.com/photo-1590490360182-c33d57733427?w=800&h=600&fit=crop",  # Minimalist room
            "https://images.unsplash.com/photo-1566665797739-1674de7a421a?w=800&h=600&fit=crop",  # Cozy single
            "https://images.unsplash.com/photo-1564501049412-61c2a3083791?w=800&h=600&fit=crop",  # Business room
            "https://images.unsplash.com/photo-1586023492125-27b2c045efd7?w=800&h=600&fit=crop",  # Comfortable single
            "https://images.unsplash.com/photo-1618221195710-dd6b41faaea8?w=800&h=600&fit=crop",  # Modern hotel room
        ],
        "double": [
            "https://images.unsplash.com/photo-1578683010236-d716f9a3f461?w=800&h=600&fit=crop",  # Luxury double
            "https://images.unsplash.com/photo-1596394516093-501ba68a0ba6?w=800&h=600&fit=crop",  # Comfortable double
            "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&h=600&fit=crop",  # Elegant double
            "https://images.unsplash.com/photo-1618221195710-dd6b41faaea8?w=800&h=600&fit=crop",  # Modern double
            "https://images.unsplash.com/photo-1631049307264-da0ec9d70304?w=800&h=600&fit=crop",  # Spacious double
            "https://images.unsplash.com/photo-1590490360182-c33d57733427?w=800&h=600&fit=crop",  # Classic double
        ],
        "family": [
            "https://images.unsplash.com/photo-1582719478250-c89cae4dc85b?w=800&h=600&fit=crop",  # Family suite
            "https://images.unsplash.com/photo-1611892440504-42a792e24d32?w=800&h=600&fit=crop",  # Large suite
            "https://images.unsplash.com/photo-1590490360182-c33d57733427?w=800&h=600&fit=crop",  # Modern suite
            "https://images.unsplash.com/photo-1566665797739-1674de7a421a?w=800&h=600&fit=crop",  # Comfortable suite
            "https://images.unsplash.com/photo-1618221195710-dd6b41faaea8?w=800&h=600&fit=crop",  # Spacious family
            "https://images.unsplash.com/photo-1578683010236-d716f9a3f461?w=800&h=600&fit=crop",  # Luxury suite
        ],
        "couple": [
            "https://images.unsplash.com/photo-1578683010236-d716f9a3f461?w=800&h=600&fit=crop",  # Honeymoon suite
            "https://images.unsplash.com/photo-1618221195710-dd6b41faaea8?w=800&h=600&fit=crop",  # Intimate room
            "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&h=600&fit=crop",  # Luxury romantic
            "https://images.unsplash.com/photo-1596394516093-501ba68a0ba6?w=800&h=600&fit=crop",  # Elegant couple
            "https://images.unsplash.com/photo-1582719478250-c89cae4dc85b?w=800&h=600&fit=crop",  # Romantic suite
            "https://images.unsplash.com/photo-1631049307264-da0ec9d70304?w=800&h=600&fit=crop",  # Premium couple
        ],
        "vip": [
            "https://images.unsplash.com/photo-1582719478250-c89cae4dc85b?w=800&h=600&fit=crop",  # Executive suite
            "https://images.unsplash.com/photo-1611892440504-42a792e24d32?w=800&h=600&fit=crop",  # Penthouse
            "https://images.unsplash.com/photo-1618221195710-dd6b41faaea8?w=800&h=600&fit=crop",  # Luxury villa
            "https://images.unsplash.com/photo-1566073771259-6a8506099945?w=800&h=600&fit=crop",  # Premium suite
            "https://images.unsplash.com/photo-1578683010236-d716f9a3f461?w=800&h=600&fit=crop",  # Presidential suite
            "https://images.unsplash.com/photo-1631049307264-da0ec9d70304?w=800&h=600&fit=crop",  # Elite residence
        ],
        "default": [
            "https://images.unsplash.com/photo-1631049307264-da0ec9d70304?w=800&h=600&fit=crop",
            "https://images.unsplash.com/photo-1590490360182-c33d57733427?w=800&h=600&fit=crop",
            "https://images.unsplash.com/photo-1566665797739-1674de7a421a?w=800&h=600&fit=crop",
        ],
    }

    # Determine room type category
    if not room_type:
        room_type = "default"

    # Normalize room type
    room_type_lower = str(room_type).lower()
    if "single" in room_type_lower or room_type_lower == "1":
        category = "single"
    elif "double" in room_type_lower or room_type_lower == "2":
        category = "double"
    elif "family" in room_type_lower or "suite" in room_type_lower:
        category = "family"
    elif (
        "couple" in room_type_lower
        or "romantic" in room_type_lower
        or "honeymoon" in room_type_lower
    ):
        category = "couple"
    elif (
        "vip" in room_type_lower
        or "presidential" in room_type_lower
        or "executive" in room_type_lower
    ):
        category = "vip"
    else:
        category = "default"

    # Get collection for this category
    images = room_image_collection.get(category, room_image_collection["default"])

    # Select image based on room index (round-robin)
    selected_image = images[room_index % len(images)]

    return selected_image


def get_360_tour_image(room_type=None, room_index=0):
    """
    Get a unique 360° panorama image URL based on room type and index.
    Returns different images for different room types to show variety.
    """
    # Collection of free 360° panorama images (equirectangular format)
    # These are demo images - replace with actual room photos in production

    panorama_collection = {
        "single": [
            "https://pannellum.org/images/alma.jpg",  # Classic room
            "https://pannellum.org/images/cerro-toco-0.jpg",  # Modern hotel room
            "https://pannellum.org/images/cerro-toco-1.jpg",  # Cozy room
            "https://pannellum.org/images/cerro-toco-2.jpg",  # Minimalist room
            "https://pannellum.org/images/cerro-toco-3.jpg",  # Business room
        ],
        "double": [
            "https://pannellum.org/images/alma.jpg",  # Standard double
            "https://pannellum.org/images/cerro-toco-0.jpg",  # Luxury double
            "https://pannellum.org/images/cerro-toco-1.jpg",  # Comfortable double
            "https://pannellum.org/images/cerro-toco-2.jpg",  # Modern double
            "https://pannellum.org/images/cerro-toco-3.jpg",  # Elegant double
        ],
        "family": [
            "https://pannellum.org/images/alma.jpg",  # Spacious suite
            "https://pannellum.org/images/cerro-toco-0.jpg",  # Family suite
            "https://pannellum.org/images/cerro-toco-1.jpg",  # Large suite
            "https://pannellum.org/images/cerro-toco-2.jpg",  # Modern suite
            "https://pannellum.org/images/cerro-toco-3.jpg",  # Comfortable suite
        ],
        "couple": [
            "https://pannellum.org/images/alma.jpg",  # Romantic room
            "https://pannellum.org/images/cerro-toco-0.jpg",  # Honeymoon suite
            "https://pannellum.org/images/cerro-toco-1.jpg",  # Intimate room
            "https://pannellum.org/images/cerro-toco-2.jpg",  # Luxury romantic
            "https://pannellum.org/images/cerro-toco-3.jpg",  # Elegant couple
        ],
        "vip": [
            "https://pannellum.org/images/alma.jpg",  # Presidential suite
            "https://pannellum.org/images/cerro-toco-0.jpg",  # Executive suite
            "https://pannellum.org/images/cerro-toco-1.jpg",  # Penthouse
            "https://pannellum.org/images/cerro-toco-2.jpg",  # Luxury villa
            "https://pannellum.org/images/cerro-toco-3.jpg",  # Premium suite
        ],
        "default": [
            "https://pannellum.org/images/alma.jpg",
            "https://pannellum.org/images/cerro-toco-0.jpg",
            "https://pannellum.org/images/cerro-toco-1.jpg",
        ],
    }

    # Determine room type category
    if not room_type:
        room_type = "default"

    # Normalize room type
    room_type_lower = str(room_type).lower()
    if "single" in room_type_lower or room_type_lower == "1":
        category = "single"
    elif "double" in room_type_lower or room_type_lower == "2":
        category = "double"
    elif "family" in room_type_lower or "suite" in room_type_lower:
        category = "family"
    elif (
        "couple" in room_type_lower
        or "romantic" in room_type_lower
        or "honeymoon" in room_type_lower
    ):
        category = "couple"
    elif (
        "vip" in room_type_lower
        or "presidential" in room_type_lower
        or "executive" in room_type_lower
    ):
        category = "vip"
    else:
        category = "default"

    # Get collection for this category
    images = panorama_collection.get(category, panorama_collection["default"])

    # Select image based on room index (round-robin)
    selected_image = images[room_index % len(images)]

    return selected_image


def hash_password(password):
    """Hash password using SHA256"""
    return hashlib.sha256(password.encode()).hexdigest()


def check_password(password, hashed_password):
    """Verify password against hash"""
    return hash_password(password) == hashed_password


def get_user_by_email(email):
    """Get user by email address"""
    try:
        user = get_user(email.lower())
        return user
    except Exception as e:
        print(f"Error retrieving user: {e}")
        return None


# Decorators for role-based access
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Please login to access this page", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def role_required(required_role):
    """Decorator to require specific role. Supports multiple roles or branch-specific roles."""

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user_role = session.get("role", "guest")
            user_branch_id = session.get("branch_id")

            # Handle multiple roles
            if isinstance(required_role, list):
                if user_role not in required_role:
                    flash("Unauthorized access", "danger")
                    return redirect(url_for("dashboard"))
            # Handle branch-specific roles
            elif required_role in ["branch_manager", "branch_staff"]:
                if user_role != required_role:
                    flash("Unauthorized access", "danger")
                    return redirect(url_for("dashboard"))
                # Check if branch_id matches (for branch-specific operations)
                if "branch_id" in kwargs and kwargs["branch_id"]:
                    if user_branch_id != kwargs["branch_id"]:
                        flash("You can only access your assigned branch", "danger")
                        return redirect(url_for("dashboard"))
            # Standard role check
            else:
                if user_role != required_role:
                    flash("Unauthorized access", "danger")
                    return redirect(url_for("dashboard"))

            return f(*args, **kwargs)

        return decorated_function

    return decorator


def branch_access_required(f):
    """Decorator to ensure user has access to the branch they're trying to access"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        user_role = session.get("role", "guest")
        user_branch_id = session.get("branch_id")

        # Super admin has access to all branches
        if user_role == "super_admin":
            return f(*args, **kwargs)

        # Branch managers and staff can only access their branch
        if user_role in ["branch_manager", "branch_staff"]:
            branch_id = kwargs.get("branch_id") or request.args.get("branch_id")
            if branch_id and user_branch_id != branch_id:
                flash("You can only access your assigned branch", "danger")
                return redirect(url_for("dashboard"))

        return f(*args, **kwargs)

    return decorated_function


def allowed_file(filename):
    """Check if file extension is allowed"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_room_image(file):
    """Save room image and return filename"""
    if file and file.filename and allowed_file(file.filename):
        try:
            # Generate unique filename
            ext = file.filename.rsplit(".", 1)[1].lower()
            filename = f"room_{uuid.uuid4()}_{int(datetime.now().timestamp())}.{ext}"

            filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
            file.save(filepath)

            print(f"Image saved: {filename}")
            return filename
        except Exception as e:
            print(f"Error saving image: {e}")
            return None
    return None


# Coupon codes (stored in memory - in production, use database)
VALID_COUPONS = {
    "SAVE10": {"discount": 10, "type": "percentage"},
    "SAVE500": {"discount": 500, "type": "fixed"},
    "WELCOME15": {"discount": 15, "type": "percentage"},
    "SUMMER20": {"discount": 20, "type": "percentage"},
}


@app.route("/language/<language>")
def set_language(language=None):
    if language in LANGUAGES:
        session["language"] = language
    return redirect(request.referrer or url_for("index"))


@app.route("/branch/<branch_id>")
def branch_detail(branch_id):
    """Branch detail page"""
    try:
        branch = get_branch(branch_id)
        if not branch:
            flash("Branch not found", "danger")
            return redirect(url_for("index"))

        # Get rooms for this branch
        branch_rooms = get_all_rooms(branch_id=branch_id)

        return render_template(
            "branch_detail.html",
            branch=branch,
            rooms=branch_rooms,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading branch: {e}")
        flash(f"Error loading branch: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/room/<room_id>/virtual-tour")
def virtual_tour(room_id):
    """Virtual 360° room tour"""
    try:
        all_rooms = get_all_rooms()
        room = next((r for r in all_rooms if r.get("room_id") == room_id), None)

        if not room:
            flash("Room not found", "danger")
            return redirect(url_for("rooms"))

        # Generate default 360° image URL if not provided
        # In production, these would be actual 360° panorama images
        if not room.get("virtual_tour_360_url"):
            # Get unique 360° image based on room type and ID hash
            room_type = room.get("room_type", "default")
            room_index = hash(room_id) % 10  # Use room ID hash for variety
            room["virtual_tour_360_url"] = get_360_tour_image(room_type, room_index)

        # Add Matterport URL if available (optional - can be set per room)
        # Format: https://my.matterport.com/show/?m=XXXXX
        if not room.get("matterport_url"):
            # Example Matterport URL (replace with actual Matterport model IDs)
            # room['matterport_url'] = 'https://my.matterport.com/show/?m=XXXXX'
            room["matterport_url"] = None  # Set to None if not configured

        # Default hotspots if not provided
        if not room.get("hotspots"):
            room["hotspots"] = [
                {
                    "pitch": -5,
                    "yaw": 0,
                    "label": "Bed",
                    "description": f"{'King-size bed' if room.get('room_type') in ['couple', 'vip'] else 'Comfortable bed'} with premium linens",
                },
                {
                    "pitch": -10,
                    "yaw": 90,
                    "label": "Bathroom",
                    "description": "Modern private bathroom with premium amenities",
                },
                {
                    "pitch": 10,
                    "yaw": 180,
                    "label": "View",
                    "description": "Stunning views from your room",
                },
            ]

        return render_template(
            "virtual_tour.html", room=room, user=session.get("user_info", {})
        )
    except Exception as e:
        print(f"Error loading virtual tour: {e}")
        flash(f"Error loading virtual tour: {str(e)}", "danger")
        return redirect(url_for("book_room", room_id=room_id))


@app.route("/")
def index():
    """Home/Landing page with branch selection"""
    featured_rooms = []
    branches = []
    try:
        if DB_AVAILABLE:
            # Get all active branches
            branches = get_all_branches(status=None)

            # Get featured rooms from all branches
            all_rooms = get_all_rooms()
            featured_rooms = [
                room for room in all_rooms if room.get("availability") == "available"
            ][:6]
        else:
            # Mock data for testing
            branches = [
                {
                    "branch_id": "BLISS-MUM",
                    "branch_name": "Blissful Abodes Mumbai",
                    "location": {"city": "Mumbai"},
                    "starting_price": 4500,
                },
                {
                    "branch_id": "BLISS-DEL",
                    "branch_name": "Blissful Abodes Delhi",
                    "location": {"city": "Delhi"},
                    "starting_price": 5000,
                },
            ]
            featured_rooms = [
                {
                    "room_id": "1",
                    "name": "Presidential Suite",
                    "location": "Mumbai, Maharashtra",
                    "price": 4500,
                    "availability": "available",
                    "branch_id": "BLISS-MUM",
                },
                {
                    "room_id": "2",
                    "name": "Ocean View Deluxe",
                    "location": "Goa",
                    "price": 6000,
                    "availability": "available",
                    "branch_id": "BLISS-GOA",
                },
            ]
    except Exception as e:
        print(f"Error loading homepage data: {e}")
        featured_rooms = []
        branches = []

    return render_template(
        "index.html",
        featured_rooms=featured_rooms,
        branches=branches,
        user=session.get("user_info"),
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    """User login"""
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_user_by_email(email)

        print(f"DEBUG: Login attempt for {email}")
        print(f"DEBUG: User found: {user is not None}")
        if user:
            print(
                f"DEBUG: Password hash match: {check_password(password, user.get('password', ''))}"
            )

        if user and check_password(password, user.get("password", "")):
            # Login successful for any role
            user_role = user.get("role", "guest")
            session["user_id"] = user.get("user_id") or user.get("id")
            session["role"] = user_role
            session["branch_id"] = user.get(
                "branch_id"
            )  # Store branch_id for branch-specific users
            session["user_info"] = {
                "name": user.get("name", ""),
                "email": user.get("email", ""),
                "role": user_role,
                "branch_id": user.get("branch_id"),
            }

            # Send SNS notification to admins about login
            try:
                from sns_notifier import send_notification
                from datetime import datetime

                subject = f"{user_role.title()} Login - Blissful Abodes"
                message = f"""
User logged in to Blissful Abodes:

Name: {user.get('name', 'Unknown')}
Email: {user.get('email', 'Unknown')}
Role: {user_role.upper()}
Login Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

This is an automated notification sent to subscribed admins.
                """
                send_notification(subject, message)
                print(f"✓ Sent login notification for {user.get('email')}")
            except Exception as e:
                print(f"Error sending login notification: {e}")

            flash(f'Welcome back, {user.get("name")}!', "success")

            # Redirect based on role
            if user_role == "super_admin":
                return redirect(url_for("super_admin_dashboard"))
            elif user_role == "admin":
                return redirect(url_for("admin_dashboard"))
            elif user_role == "manager":
                return redirect(url_for("manager_dashboard"))
            elif user_role == "staff":
                return redirect(url_for("staff_dashboard"))
            else:  # guest and any other role
                return redirect(url_for("dashboard"))
        else:
            flash("Invalid email or password", "danger")
            print(f"DEBUG: Login failed for {email}")

    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    """User registration - Guest only"""
    if request.method == "POST":
        try:
            user_data = {
                "user_id": str(uuid.uuid4()),
                "name": request.form.get("name", "").strip(),
                "email": request.form.get("email", "").strip().lower(),
                "password": hash_password(request.form.get("password", "")),
                "age": int(request.form.get("age", 0)),
                "role": "guest",  # Always guest - NEVER allow staff/admin registration
                "created_at": datetime.now().isoformat(),
            }

            # Validate required fields
            if (
                not user_data["name"]
                or not user_data["email"]
                or not user_data["password"]
            ):
                flash("All fields are required", "danger")
                return redirect(url_for("register"))

            if user_data["age"] < 18:
                flash("You must be at least 18 years old", "danger")
                return redirect(url_for("register"))

            # Verify email doesn't already exist
            existing_user = get_user_by_email(user_data["email"])
            if existing_user:
                flash(
                    "Email already registered. Please login or use a different email.",
                    "warning",
                )
                return redirect(url_for("register"))

            # Add user to database
            add_user(user_data)

            # Send SNS notification to admins about new registration
            try:
                from sns_notifier import send_notification

                subject = (
                    f"New {user_data['role'].title()} Registration - Blissful Abodes"
                )
                message = f"""
New user registered on Blissful Abodes:

Name: {user_data['name']}
Email: {user_data['email']}
Role: {user_data['role'].upper()}
Age: {user_data['age']}
Registration Time: {user_data['created_at']}

This is an automated notification sent to subscribed admins.
                """
                send_notification(subject, message)
                print(f"✓ Sent registration notification for {user_data['email']}")
            except Exception as e:
                print(f"Error sending registration notification: {e}")

            flash("Registration successful! You can now login.", "success")

            # Return to login page
            return redirect(url_for("login"))

        except ValueError as e:
            flash(f"Invalid age value: {str(e)}", "danger")
        except Exception as e:
            flash(f"Registration error: {str(e)}", "danger")

    return render_template("register.html")


@app.route("/logout")
def logout():
    """User logout"""
    session.clear()
    flash("You have been logged out", "info")
    return redirect(url_for("index"))


@app.route("/dashboard")
@login_required
def dashboard():
    """User dashboard based on role - routes to appropriate dashboard"""
    role = session.get("role", "guest")

    if role == "super_admin":
        return redirect(url_for("super_admin_dashboard"))
    elif role == "admin":
        return redirect(url_for("admin_dashboard"))
    elif role == "manager":
        return redirect(url_for("manager_dashboard"))
    elif role == "staff":
        return redirect(url_for("staff_dashboard"))
    else:  # guest
        return redirect(url_for("guest_dashboard"))


@app.route("/dashboard/guest")
@login_required
def guest_dashboard():
    """Enhanced Guest dashboard with comprehensive stats"""
    try:
        user_id = session.get("user_id")
        if not user_id:
            return redirect(url_for("login"))

        # Get comprehensive booking data with error handling
        all_bookings = []
        upcoming_bookings = []
        past_bookings = []
        booking_stats = {
            "total_bookings": 0,
            "upcoming": 0,
            "past": 0,
            "completed": 0,
            "cancelled": 0,
            "total_spent": 0,
            "total_nights": 0,
        }
        loyalty_points = 0
        loyalty_tier = "Silver"
        user_reviews = []
        favorite_rooms = []

        try:
            all_bookings = get_user_bookings(user_id)
            print(f"DEBUG: Retrieved {len(all_bookings)} total bookings")
        except Exception as e:
            print(f"Error getting user bookings: {e}")

        try:
            upcoming_bookings = get_upcoming_bookings(user_id)
            print(f"DEBUG: Retrieved {len(upcoming_bookings)} upcoming bookings")
        except Exception as e:
            print(f"Error getting upcoming bookings: {e}")

        try:
            past_bookings = get_past_bookings(user_id)
            print(f"DEBUG: Retrieved {len(past_bookings)} past bookings")
        except Exception as e:
            print(f"Error getting past bookings: {e}")

        try:
            booking_stats = get_booking_stats(user_id)
            print(f"DEBUG: Stats - Spent: ₹{booking_stats.get('total_spent', 0)}")
        except Exception as e:
            print(f"Error getting booking stats: {e}")
            import traceback

            traceback.print_exc()

        try:
            loyalty_points, loyalty_tier = get_loyalty_points(user_id)
            print(f"DEBUG: Loyalty - {loyalty_points} points, {loyalty_tier} tier")
        except Exception as e:
            print(f"Error getting loyalty points: {e}")

        try:
            user_reviews = get_user_reviews(user_id)
            print(f"DEBUG: Retrieved {len(user_reviews)} reviews")
        except Exception as e:
            print(f"Error getting user reviews: {e}")

        try:
            favorite_rooms = get_user_favorites(user_id)
            print(f"DEBUG: Retrieved {len(favorite_rooms)} favorite rooms")
        except Exception as e:
            print(f"Error getting favorite rooms: {e}")

        return render_template(
            "dashboard.html",
            bookings=all_bookings,
            upcoming_bookings=upcoming_bookings,
            past_bookings=past_bookings,
            stats=booking_stats,
            loyalty_points=loyalty_points,
            loyalty_tier=loyalty_tier,
            reviews=user_reviews,
            favorite_rooms=favorite_rooms,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading dashboard: {str(e)}", "danger")
        return render_template(
            "dashboard.html",
            bookings=[],
            upcoming_bookings=[],
            past_bookings=[],
            stats={
                "total_bookings": 0,
                "upcoming": 0,
                "past": 0,
                "completed": 0,
                "cancelled": 0,
                "total_spent": 0,
                "total_nights": 0,
            },
            loyalty_points=0,
            loyalty_tier="Silver",
            reviews=[],
            favorite_rooms=[],
            user=session.get("user_info", {}),
        )


@app.route("/guest/favorites/add/<room_id>", methods=["POST"])
@login_required
def add_to_favorites(room_id):
    """Add a room to user's favorites"""
    try:
        user_id = session.get("user_id")
        if not user_id:
            return jsonify({"success": False, "error": "Not logged in"})

        result = add_favorite_room(user_id, room_id)
        return jsonify({"success": True, "message": "Added to favorites"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/guest/favorites/remove/<room_id>", methods=["POST"])
@login_required
def remove_from_favorites(room_id):
    """Remove a room from user's favorites"""
    try:
        user_id = session.get("user_id")
        if not user_id:
            return jsonify({"success": False, "error": "Not logged in"})

        result = remove_favorite_room(user_id, room_id)
        return jsonify({"success": True, "message": "Removed from favorites"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/guest/account-settings", methods=["GET", "POST"])
@login_required
def account_settings():
    """Guest account settings page"""
    try:
        user_id = session.get("user_id")
        if not user_id:
            return redirect(url_for("login"))

        user = get_user(user_id)

        if request.method == "POST":
            # Update user profile
            updated_data = {
                "user_id": user_id,
                "name": request.form.get("name", user.get("name")),
                "email": request.form.get("email", user.get("email")),
                "phone": request.form.get("phone", user.get("phone", "")),
                "age": int(request.form.get("age", user.get("age", 18))),
                "role": user.get("role", "guest"),  # Keep existing role
            }

            # Update password if provided
            new_password = request.form.get("new_password", "").strip()
            if new_password:
                updated_data["password"] = hash_password(new_password)
            else:
                updated_data["password"] = user.get("password")

            # Validate email uniqueness (if changed)
            if updated_data["email"] != user.get("email"):
                existing = get_user(updated_data["email"])
                if existing and existing.get("user_id") != user_id:
                    flash("Email already in use by another account", "danger")
                    return redirect(url_for("account_settings"))

            update_user(updated_data)

            # Update session
            session["user_info"] = {
                "name": updated_data["name"],
                "email": updated_data["email"],
                "age": updated_data["age"],
                "phone": updated_data.get("phone", ""),
            }

            flash("Account settings updated successfully!", "success")
            return redirect(url_for("account_settings"))

        return render_template(
            "account_settings.html",
            user=user,
            session_user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error in account settings: {e}")
        flash(f"Error: {str(e)}", "danger")
        return redirect(url_for("guest_dashboard"))


@app.route("/dashboard/staff")
@login_required
@role_required(["staff", "manager", "admin", "super_admin"])
def staff_dashboard():
    """Staff dashboard"""
    try:
        bookings = get_all_bookings()
        rooms = get_all_rooms()

        # Get chatbot requests for staff (reports and service requests)
        all_chat_requests = get_chat_requests()
        pending_requests = [
            r for r in all_chat_requests if r.get("status") == "pending"
        ]
        service_requests = [
            r for r in all_chat_requests if r.get("request_type") == "extra_service"
        ]
        report_requests = [
            r for r in all_chat_requests if r.get("request_type") == "report"
        ]

        return render_template(
            "staff_dashboard.html",
            bookings=bookings,
            rooms=rooms,
            pending_requests=pending_requests,
            service_requests=service_requests,
            report_requests=report_requests,
            user=session.get("user_info", {}),
            now=datetime.now(),
        )
    except Exception as e:
        print(f"Error loading staff dashboard: {e}")
        flash(f"Error loading staff dashboard: {str(e)}", "danger")
        return render_template(
            "staff_dashboard.html",
            bookings=[],
            rooms=[],
            pending_requests=[],
            service_requests=[],
            report_requests=[],
            user=session.get("user_info", {}),
            now=datetime.now(),
        )


@app.route("/dashboard/manager")
@login_required
@role_required(["manager", "admin", "super_admin"])
def manager_dashboard():
    """Manager Dashboard for Department Heads"""
    try:
        all_rooms = get_all_rooms()
        all_bookings = get_all_bookings()
        all_users = get_all_users()
        all_reviews = get_all_reviews()

        # Staff members (staff role only)
        staff_members = [u for u in all_users if u.get("role") == "staff"]

        # Room stats
        total_rooms = len(all_rooms)
        occupied_rooms = len(
            [
                r
                for r in all_rooms
                if r.get("availability") in ["occupied", "unavailable"]
            ]
        )
        available_rooms = total_rooms - occupied_rooms
        occupancy_rate = round(
            (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0, 1
        )

        # Today's activity
        today = datetime.now().date()
        todays_checkins = [
            b
            for b in all_bookings
            if b.get("check_in") and str(b.get("check_in", ""))[:10] == str(today)
        ]
        todays_checkouts = [
            b
            for b in all_bookings
            if b.get("check_out") and str(b.get("check_out", ""))[:10] == str(today)
        ]

        # Today revenue (confirmed bookings today)
        today_revenue = sum(
            float(b.get("total_price", 0))
            for b in todays_checkins
            if b.get("status") == "confirmed"
        )

        # Guest satisfaction
        avg_rating = 0
        if all_reviews:
            avg_rating = round(
                sum(float(r.get("rating", 0)) for r in all_reviews) / len(all_reviews),
                1,
            )

        # Pending requests (chat requests)
        all_requests = []
        try:
            all_requests = get_chat_requests()
        except Exception:
            pass
        pending_requests = [r for r in all_requests if r.get("status") == "pending"]

        # Mock housekeeping tasks
        housekeeping_tasks = [
            {
                "room": "S100",
                "task": "Deep Clean",
                "assigned_to": "Priya",
                "priority": "High",
                "status": "In Progress",
            },
            {
                "room": "D130",
                "task": "Room Clean",
                "assigned_to": "Lakshmi",
                "priority": "Medium",
                "status": "Pending",
            },
            {
                "room": "V185",
                "task": "Inspection",
                "assigned_to": "Meena",
                "priority": "Low",
                "status": "Scheduled",
            },
            {
                "room": "F160",
                "task": "Turndown",
                "assigned_to": "Ravi",
                "priority": "High",
                "status": "Pending",
            },
        ]

        # Mock inventory
        inventory = [
            {"item": "Towels", "stock": 45, "reorder": 50, "status": "low"},
            {"item": "Bed Sheets", "stock": 120, "reorder": 80, "status": "ok"},
            {"item": "Toiletries", "stock": 200, "reorder": 100, "status": "ok"},
            {"item": "Hangers", "stock": 20, "reorder": 50, "status": "critical"},
            {"item": "Key Cards", "stock": 300, "reorder": 200, "status": "ok"},
        ]

        # Mock pending approvals
        pending_approvals = [
            {
                "type": "Discount Request",
                "detail": "15% off for corporate booking",
                "amount": "₹18,000 → ₹15,300",
                "from": "Front Desk",
            },
            {
                "type": "Leave Request",
                "detail": "Priya Sharma - Feb 15-16 (2 days)",
                "reason": "Family function",
                "from": "Priya Sharma",
            },
            {
                "type": "Expense Claim",
                "detail": "Guest amenities",
                "amount": "₹2,500",
                "from": "Raj Kumar",
            },
        ]

        return render_template(
            "manager_dashboard.html",
            staff_members=staff_members,
            all_rooms=all_rooms,
            all_bookings=all_bookings,
            all_reviews=all_reviews,
            total_rooms=total_rooms,
            occupied_rooms=occupied_rooms,
            available_rooms=available_rooms,
            occupancy_rate=occupancy_rate,
            today_revenue=today_revenue,
            avg_rating=avg_rating,
            todays_checkins=todays_checkins,
            todays_checkouts=todays_checkouts,
            pending_requests=pending_requests,
            housekeeping_tasks=housekeeping_tasks,
            inventory=inventory,
            pending_approvals=pending_approvals,
            user=session.get("user_info", {}),
            now=datetime.now(),
        )
    except Exception as e:
        print(f"Error loading manager dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading dashboard: {str(e)}", "danger")
        return render_template(
            "manager_dashboard.html",
            staff_members=[],
            all_rooms=[],
            all_bookings=[],
            all_reviews=[],
            total_rooms=0,
            occupied_rooms=0,
            available_rooms=0,
            occupancy_rate=0,
            today_revenue=0,
            avg_rating=0,
            todays_checkins=[],
            todays_checkouts=[],
            pending_requests=[],
            housekeeping_tasks=[],
            inventory=[],
            pending_approvals=[],
            user=session.get("user_info", {}),
            now=datetime.now(),
        )


@app.route("/dashboard/super-admin")
@login_required
@role_required("super_admin")
def super_admin_dashboard():
    """Super Admin Dashboard — Executive Command Center"""
    try:
        branches = get_all_branches(status="active")
        analytics = get_analytics()
        all_rooms = get_all_rooms()
        all_bookings = get_all_bookings()
        all_users = get_all_users()
        all_reviews = get_all_reviews()

        # Compute key metrics
        total_rooms = len(all_rooms)
        occupied_rooms = len(
            [
                r
                for r in all_rooms
                if r.get("availability") in ["occupied", "unavailable"]
            ]
        )
        occupancy_rate = round(
            (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0, 1
        )
        total_revenue = sum(float(b.get("total_price", 0)) for b in all_bookings)

        today = datetime.now()
        current_month = today.strftime("%Y-%m")
        monthly_revenue = sum(
            float(b.get("total_price", 0))
            for b in all_bookings
            if str(b.get("check_in", ""))[:7] == current_month
        )
        if monthly_revenue == 0:
            monthly_revenue = total_revenue * 0.08 if total_revenue > 0 else 150000

        avg_rating = 4.9
        if all_reviews:
            avg_rating = round(
                sum(float(r.get("rating", 0)) for r in all_reviews) / len(all_reviews),
                1,
            )

        # Branch analytics
        branch_analytics = {}
        for branch in branches:
            branch_id = branch.get("branch_id")
            branch_rooms = get_all_rooms(branch_id=branch_id)
            branch_bookings = [
                b for b in all_bookings if b.get("branch_id") == branch_id
            ]
            br_total = len(branch_rooms)
            br_occ = len(
                [r for r in branch_rooms if r.get("availability") == "unavailable"]
            )
            br_rev = sum(
                float(b.get("total_price", 0))
                for b in branch_bookings
                if b.get("payment_status") == "paid"
            )
            branch_analytics[branch_id] = {
                "branch_name": branch.get("branch_name", ""),
                "total_rooms": br_total,
                "occupied_rooms": br_occ,
                "occupancy_rate": round(
                    (br_occ / br_total * 100) if br_total > 0 else 0, 1
                ),
                "total_bookings": len(branch_bookings),
                "total_revenue": br_rev,
                "avg_rating": 4.5,
            }

        # Chat requests
        try:
            all_chat_requests = get_chat_requests()
            pending_requests = [
                r for r in all_chat_requests if r.get("status") == "pending"
            ]
            service_requests = [
                r for r in all_chat_requests if r.get("request_type") == "extra_service"
            ]
            report_requests = [
                r for r in all_chat_requests if r.get("request_type") == "report"
            ]
        except Exception:
            pending_requests, service_requests, report_requests = [], [], []

        return render_template(
            "super_admin_dashboard.html",
            analytics=analytics,
            branch_analytics=branch_analytics,
            branches=branches,
            all_rooms=all_rooms,
            all_bookings=all_bookings,
            all_users=all_users,
            all_reviews=all_reviews,
            rooms=all_rooms,
            bookings=all_bookings,
            total_rooms=total_rooms,
            occupied_rooms=occupied_rooms,
            occupancy_rate=occupancy_rate,
            total_revenue=total_revenue,
            monthly_revenue=monthly_revenue,
            avg_rating=avg_rating,
            pending_requests=pending_requests,
            service_requests=service_requests,
            report_requests=report_requests,
            user=session.get("user_info", {}),
            now=today,
        )
    except Exception as e:
        print(f"Error loading super admin dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading dashboard: {str(e)}", "danger")
        return render_template(
            "super_admin_dashboard.html",
            analytics={},
            branch_analytics={},
            branches=[],
            all_rooms=[],
            all_bookings=[],
            all_users=[],
            all_reviews=[],
            rooms=[],
            bookings=[],
            total_rooms=100,
            occupied_rooms=0,
            occupancy_rate=0,
            total_revenue=0,
            monthly_revenue=0,
            avg_rating=4.9,
            pending_requests=[],
            service_requests=[],
            report_requests=[],
            user=session.get("user_info", {}),
            now=datetime.now(),
        )


@app.route("/dashboard/branch-manager")
@login_required
@role_required("branch_manager")
@branch_access_required
def branch_manager_dashboard():
    """Branch Manager Dashboard - Branch-specific view"""
    try:
        branch_id = session.get("branch_id")
        if not branch_id:
            flash("No branch assigned to your account", "danger")
            return redirect(url_for("index"))

        branch = get_branch(branch_id)
        if not branch:
            flash("Branch not found", "danger")
            return redirect(url_for("index"))

        # Get branch-specific data
        branch_rooms = get_all_rooms(branch_id=branch_id)
        branch_bookings = [
            b for b in get_all_bookings() if b.get("branch_id") == branch_id
        ]

        # Today's operations
        today = datetime.now().date()
        today_check_ins = [
            b for b in branch_bookings if b.get("check_in") == today.isoformat()
        ]
        today_check_outs = [
            b for b in branch_bookings if b.get("check_out") == today.isoformat()
        ]
        today_revenue = sum(
            float(b.get("total_price", 0))
            for b in branch_bookings
            if b.get("check_in") == today.isoformat()
            and b.get("payment_status") == "paid"
        )

        # Occupancy
        total_rooms = len(branch_rooms)
        occupied_rooms = len(
            [r for r in branch_rooms if r.get("availability") == "unavailable"]
        )
        occupancy_rate = (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0

        # Room type performance
        room_type_performance = {}
        for room in branch_rooms:
            room_type = room.get("room_type", "unknown")
            if room_type not in room_type_performance:
                room_type_performance[room_type] = {"total": 0, "occupied": 0}
            room_type_performance[room_type]["total"] += 1
            if room.get("availability") == "unavailable":
                room_type_performance[room_type]["occupied"] += 1

        # Calculate occupancy per room type
        for room_type in room_type_performance:
            total = room_type_performance[room_type]["total"]
            occupied = room_type_performance[room_type]["occupied"]
            room_type_performance[room_type]["occupancy_rate"] = (
                (occupied / total * 100) if total > 0 else 0
            )

        # Get chatbot requests for this branch
        all_chat_requests = get_chat_requests()
        branch_requests = [
            r for r in all_chat_requests if r.get("branch_id") == branch_id
        ]
        pending_requests = [r for r in branch_requests if r.get("status") == "pending"]
        service_requests = [
            r for r in branch_requests if r.get("request_type") == "extra_service"
        ]
        report_requests = [
            r for r in branch_requests if r.get("request_type") == "report"
        ]

        return render_template(
            "branch_manager_dashboard.html",
            branch=branch,
            branch_rooms=branch_rooms,
            branch_bookings=branch_bookings,
            today_check_ins=today_check_ins,
            today_check_outs=today_check_outs,
            today_revenue=today_revenue,
            occupancy_rate=occupancy_rate,
            room_type_performance=room_type_performance,
            pending_requests=pending_requests,
            service_requests=service_requests,
            report_requests=report_requests,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading branch manager dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading dashboard: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/dashboard/branch-staff")
@login_required
@role_required("branch_staff")
@branch_access_required
def branch_staff_dashboard():
    """Branch Staff Dashboard - Branch-specific view for staff"""
    try:
        branch_id = session.get("branch_id")
        if not branch_id:
            flash("No branch assigned to your account", "danger")
            return redirect(url_for("index"))

        branch = get_branch(branch_id)
        if not branch:
            flash("Branch not found", "danger")
            return redirect(url_for("index"))

        # Get branch-specific bookings and rooms
        branch_bookings = [
            b for b in get_all_bookings() if b.get("branch_id") == branch_id
        ]
        branch_rooms = get_all_rooms(branch_id=branch_id)

        # Today's operations
        today = datetime.now().date()
        today_check_ins = [
            b for b in branch_bookings if b.get("check_in") == today.isoformat()
        ]
        today_check_outs = [
            b for b in branch_bookings if b.get("check_out") == today.isoformat()
        ]

        # Upcoming check-ins (next 3 days)
        upcoming_dates = [(today + timedelta(days=i)).isoformat() for i in range(1, 4)]
        upcoming_check_ins = [
            b for b in branch_bookings if b.get("check_in") in upcoming_dates
        ]

        # Get chatbot requests for this branch
        all_chat_requests = get_chat_requests()
        branch_requests = [
            r for r in all_chat_requests if r.get("branch_id") == branch_id
        ]
        pending_requests = [r for r in branch_requests if r.get("status") == "pending"]
        service_requests = [
            r for r in branch_requests if r.get("request_type") == "extra_service"
        ]
        report_requests = [
            r for r in branch_requests if r.get("request_type") == "report"
        ]

        return render_template(
            "branch_staff_dashboard.html",
            branch=branch,
            branch_bookings=branch_bookings,
            branch_rooms=branch_rooms,
            today_check_ins=today_check_ins,
            today_check_outs=today_check_outs,
            upcoming_check_ins=upcoming_check_ins,
            pending_requests=pending_requests,
            service_requests=service_requests,
            report_requests=report_requests,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading branch staff dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading dashboard: {str(e)}", "danger")
        return redirect(url_for("index"))


@app.route("/dashboard/admin")
@login_required
@role_required("admin")
def admin_dashboard():
    """Admin Dashboard — Hotel Control Center"""
    try:
        all_rooms = get_all_rooms()
        all_bookings = get_all_bookings()
        all_users = get_all_users()
        all_reviews = get_all_reviews()

        # Room KPIs
        total_rooms = len(all_rooms)
        occupied_rooms = len(
            [
                r
                for r in all_rooms
                if r.get("availability") in ["occupied", "unavailable"]
            ]
        )
        occupancy_rate = round(
            (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0, 1
        )

        # Revenue KPIs
        total_revenue = sum(float(b.get("total_price", 0)) for b in all_bookings)
        today = datetime.now()
        current_month = today.strftime("%Y-%m")
        monthly_revenue = sum(
            float(b.get("total_price", 0))
            for b in all_bookings
            if str(b.get("check_in", ""))[:7] == current_month
        )
        if monthly_revenue == 0:
            monthly_revenue = total_revenue * 0.08 if total_revenue > 0 else 150000

        # ADR and RevPAR
        confirmed = [
            b for b in all_bookings if b.get("status") in ["confirmed", "completed"]
        ]
        nights_total = sum(int(b.get("nights", 1)) for b in confirmed) or 1
        confirmed_rev = sum(float(b.get("total_price", 0)) for b in confirmed)
        adr = round(confirmed_rev / nights_total) if nights_total > 0 else 4500
        revpar = round(adr * occupancy_rate / 100) if adr > 0 else 0

        # Revenue target (120% of last month estimate)
        revenue_target = monthly_revenue * 1.2 if monthly_revenue > 0 else 180000
        revenue_target_pct = min(
            round(monthly_revenue / revenue_target * 100) if revenue_target > 0 else 0,
            100,
        )

        # Guest satisfaction
        avg_rating = 0.0
        if all_reviews:
            avg_rating = round(
                sum(float(r.get("rating", 0)) for r in all_reviews) / len(all_reviews),
                1,
            )
        if avg_rating == 0:
            avg_rating = 4.7

        # Profit margin
        profit_margin = 24

        # Total users
        total_users = len(all_users)

        # Pending requests
        alerts = []
        try:
            all_chat_requests = get_chat_requests()
            pending_requests = [
                r for r in all_chat_requests if r.get("status") == "pending"
            ]
        except Exception:
            pending_requests = []

        return render_template(
            "admin_dashboard.html",
            all_rooms=all_rooms,
            all_bookings=all_bookings,
            all_users=all_users,
            all_reviews=all_reviews,
            total_rooms=total_rooms,
            occupied_rooms=occupied_rooms,
            occupancy_rate=occupancy_rate,
            monthly_revenue=monthly_revenue,
            total_revenue=total_revenue,
            revenue_target=revenue_target,
            revenue_target_pct=revenue_target_pct,
            adr=adr,
            revpar=revpar,
            avg_rating=avg_rating,
            profit_margin=profit_margin,
            total_users=total_users,
            pending_requests=pending_requests,
            alerts=alerts,
            user=session.get("user_info", {}),
            now=today,
        )
    except Exception as e:
        print(f"Error loading admin dashboard: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading admin dashboard: {str(e)}", "danger")
        return render_template(
            "admin_dashboard.html",
            all_rooms=[],
            all_bookings=[],
            all_users=[],
            all_reviews=[],
            total_rooms=100,
            occupied_rooms=0,
            occupancy_rate=0,
            monthly_revenue=0,
            total_revenue=0,
            revenue_target=180000,
            revenue_target_pct=0,
            adr=0,
            revpar=0,
            avg_rating=4.7,
            profit_margin=24,
            total_users=0,
            pending_requests=[],
            alerts=[],
            user=session.get("user_info", {}),
            now=datetime.now(),
        )


@app.route("/rooms")
def rooms():
    """Browse available rooms - branch-specific"""
    try:
        branch_id = request.args.get("branch_id", "")
        location = request.args.get("location", "")
        check_in = request.args.get("check_in", "")
        check_out = request.args.get("check_out", "")
        room_type = request.args.get("room_type", "")
        guests = request.args.get("guests", "1")

        # Get all branches for filter dropdown
        branches = get_all_branches(status=None)

        # Get all rooms first (don't filter by availability - show all)
        if branch_id:
            # Filter by branch_id - also include rooms without branch_id for backward compatibility
            all_rooms_with_branch = get_all_rooms(branch_id=branch_id)
            # Also get rooms that match the branch city/location
            branch = get_branch(branch_id)
            branch_city = ""
            if branch:
                branch_city = (
                    branch.get("location", {}).get("city", "")
                    if isinstance(branch.get("location"), dict)
                    else branch.get("location", {}).get("city", "")
                )

            # Get all rooms and filter by branch_id or location match
            all_rooms = get_all_rooms()
            filtered_rooms = []
            for room in all_rooms:
                room_branch_id = room.get("branch_id")

                # Only include if branch_id exactly matches
                if room_branch_id == branch_id:
                    filtered_rooms.append(room)

            all_rooms = filtered_rooms
            print(
                f"DEBUG: Found {len(all_rooms)} rooms for branch {branch_id} (city: {branch_city})"
            )
        elif location:
            # Filter by location
            all_rooms = get_rooms(
                location=location, availability=None
            )  # Get all, not just available
        else:
            all_rooms = get_all_rooms()

        # Filter by room type if provided
        if room_type:
            all_rooms = [r for r in all_rooms if r.get("room_type") == room_type]

        # Filter by capacity (guests)
        if guests:
            try:
                guest_count = int(guests)
                all_rooms = [
                    r for r in all_rooms if r.get("capacity", 1) >= guest_count
                ]
            except ValueError:
                pass

        # Filter by date availability if check_in and check_out provided
        if check_in and check_out:
            try:
                # Convert date format if needed (handle DD-MM-YYYY or YYYY-MM-DD)
                if "-" in check_in and len(check_in.split("-")[0]) == 2:
                    # DD-MM-YYYY format
                    check_in_parts = check_in.split("-")
                    check_in = (
                        f"{check_in_parts[2]}-{check_in_parts[1]}-{check_in_parts[0]}"
                    )
                if "-" in check_out and len(check_out.split("-")[0]) == 2:
                    # DD-MM-YYYY format
                    check_out_parts = check_out.split("-")
                    check_out = f"{check_out_parts[2]}-{check_out_parts[1]}-{check_out_parts[0]}"

                check_in_date = datetime.strptime(check_in, "%Y-%m-%d").date()
                check_out_date = datetime.strptime(check_out, "%Y-%m-%d").date()

                # Get all bookings for date range
                all_bookings = get_all_bookings()
                booked_room_ids = set()

                for booking in all_bookings:
                    if (
                        booking.get("booking_status") == "confirmed"
                        or booking.get("status") == "confirmed"
                    ):
                        try:
                            booking_check_in = datetime.strptime(
                                booking.get("check_in", ""), "%Y-%m-%d"
                            ).date()
                            booking_check_out = datetime.strptime(
                                booking.get("check_out", ""), "%Y-%m-%d"
                            ).date()

                            # Check if dates overlap
                            if not (
                                check_out_date <= booking_check_in
                                or check_in_date >= booking_check_out
                            ):
                                booked_room_ids.add(booking.get("room_id"))
                        except (ValueError, TypeError):
                            continue

                # Filter out booked rooms
                all_rooms = [
                    r for r in all_rooms if r.get("room_id") not in booked_room_ids
                ]
            except (ValueError, TypeError) as e:
                print(f"Error parsing dates: {e}")
                # Continue without date filtering if date parsing fails

        # Filter to show only available rooms (unless searching all)
        # For search results, show available rooms
        rooms_list = [
            r
            for r in all_rooms
            if r.get("availability") == "available"
            or r.get("availability") == "Available"
        ]

        # If no available rooms found, show all rooms (for debugging)
        if not rooms_list and all_rooms:
            print(
                f"WARNING: No available rooms found, but {len(all_rooms)} total rooms exist"
            )
            print(
                f"Sample room availability: {[r.get('availability') for r in all_rooms[:5]]}"
            )
            # Show all rooms for debugging
            rooms_list = all_rooms

        # Ensure all rooms have image field
        for room in rooms_list:
            if not room.get("image") or room.get("image") == "default-room.jpg":
                room["image"] = ""
            # Convert Decimal price to float
            if "price" in room:
                room["price"] = convert_decimal_to_float(room["price"])

            # SCARCITY INDEX CALCULATIONS (Feature 1)
            # Create a deterministic mock scarcity value based on room ID so it stays consistent per page load
            import hashlib

            room_hash = int(
                hashlib.md5(str(room.get("room_id", "")).encode()).hexdigest(), 16
            )

            # Rooms left: Only show badge if < 5 rooms
            # Generate deterministic mock of rooms left (1 to 10)
            room["rooms_left"] = (room_hash % 10) + 1

            # People currently viewing
            # High demand = more people viewing (3 to 25)
            room["viewing_count"] = (room_hash % 22) + 3

            # Apply dynamic pricing to base price if check in/out selected
            if check_in and check_out:
                room["dynamic_price"] = calculate_dynamic_price(
                    room["price"],
                    check_in,
                    check_out,
                    room.get("room_id", ""),
                    room.get("branch_id", ""),
                )
                if room["dynamic_price"] > room["price"]:
                    room["price_surge"] = True

        # Get branch info if branch_id provided
        selected_branch = None
        if branch_id:
            selected_branch = get_branch(branch_id)

        print(f"DEBUG: Returning {len(rooms_list)} rooms after filtering")

        return render_template(
            "room_list.html",
            rooms=rooms_list,
            branches=branches,
            selected_branch=selected_branch,
            filters={
                "branch_id": branch_id,
                "location": location,
                "check_in": check_in,
                "check_out": check_out,
                "room_type": room_type,
                "guests": guests,
            },
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading rooms: {e}")
        import traceback

        traceback.print_exc()
        flash(f"Error loading rooms: {str(e)}", "danger")
        return render_template(
            "room_list.html",
            rooms=[],
            branches=[],
            filters={},
            user=session.get("user_info", {}),
        )


@app.route("/book/<room_id>", methods=["GET", "POST"])
@login_required
def book_room(room_id):
    """Book a specific room"""
    if request.method == "POST":
        try:
            if "user_id" not in session or "user_info" not in session:
                flash("Please login to book a room", "warning")
                return redirect(url_for("login"))

            user_info = session.get("user_info", {})

            # Only guests can book rooms - all other roles cannot
            user_role = user_info.get("role", "")
            if user_role != "guest":
                flash(
                    f"{user_role.title()} users cannot book rooms. Only guests can make bookings.",
                    "warning",
                )
                return redirect(url_for("rooms"))

            if not user_info.get("name") or not user_info.get("email"):
                flash("Please complete your profile before booking", "warning")
                return redirect(url_for("login"))

            # Get room to extract branch_id
            all_rooms = get_all_rooms()
            room = next((r for r in all_rooms if r.get("room_id") == room_id), None)
            if not room:
                flash("Room not found", "danger")
                return redirect(url_for("rooms"))

            booking_data = {
                "booking_id": str(uuid.uuid4()),
                "user_id": session["user_id"],
                "room_id": room_id,
                "branch_id": room.get("branch_id", ""),
                "check_in": request.form.get("check_in", "").strip(),
                "check_out": request.form.get("check_out", "").strip(),
                "nights": 0,  # Will calculate
                "total_price": 0,  # Will calculate
                "currency": "INR",
                "payment_status": "pending",
                "booking_status": "confirmed",
                "status": "confirmed",
                "created_at": datetime.now().isoformat(),
                "guest_name": user_info.get("name", ""),
                "guest_email": user_info.get("email", ""),
            }

            # Calculate nights and total price with dynamic pricing
            try:
                check_in_date = datetime.strptime(booking_data["check_in"], "%Y-%m-%d")
                check_out_date = datetime.strptime(
                    booking_data["check_out"], "%Y-%m-%d"
                )
                nights = (check_out_date - check_in_date).days
                booking_data["nights"] = nights

                # Use dynamic pricing engine
                # Convert Decimal to float for calculations
                room_price = room.get("price", 0)
                base_price = float(room_price) if room_price else 0.0
                dynamic_price = calculate_dynamic_price(
                    base_price,
                    booking_data["check_in"],
                    booking_data["check_out"],
                    room_id,
                    room.get("branch_id", ""),
                )

                # Calculate subtotal (room price * nights)
                subtotal = dynamic_price

                # Calculate service fee (10% of subtotal)
                service_fee = subtotal * 0.10

                # Calculate GST (18% of subtotal + service fee)
                gst = (subtotal + service_fee) * 0.18

                # Calculate final total
                total_with_fees = subtotal + service_fee + gst

                # Store all pricing details
                booking_data["base_price"] = Decimal(str(base_price * nights))
                booking_data["subtotal"] = Decimal(str(subtotal))
                booking_data["service_fee"] = Decimal(str(service_fee))
                booking_data["gst"] = Decimal(str(gst))
                booking_data["total_price"] = Decimal(str(total_with_fees))
                booking_data["pricing_applied"] = True

                print(
                    f"Pricing: Base={base_price * nights}, Subtotal={subtotal}, Service Fee={service_fee}, GST={gst}, Total={total_with_fees}, Nights={nights}"
                )
            except Exception as e:
                print(f"Error calculating booking price: {e}")
                # Fallback to basic pricing
                booking_data["total_price"] = Decimal(
                    str(float(room.get("price", 0)) * nights)
                )
                booking_data["pricing_applied"] = False

            # Validate dates
            if not booking_data["check_in"] or not booking_data["check_out"]:
                flash("Please select check-in and check-out dates", "danger")
                return redirect(url_for("book_room", room_id=room_id))

            # Add booking to database
            add_booking(booking_data)

            # Update room availability
            update_room_availability(room_id, "unavailable")

            # Add loyalty points (₹100 = 1 point)
            try:
                points_earned = int(booking_data["total_price"] / 100)
                if points_earned > 0:
                    add_loyalty_points(
                        session["user_id"], points_earned, reason="booking"
                    )
                    # First booking bonus
                    user_bookings = get_user_bookings(session["user_id"])
                    if len(user_bookings) == 1:
                        add_loyalty_points(
                            session["user_id"], 500, reason="first_booking_bonus"
                        )
            except Exception as e:
                print(f"Error adding loyalty points: {e}")

            # Send confirmation email and schedule reminders
            user_email = session["user_info"].get("email")
            try:
                print(f"Attempting to send booking confirmation to {user_email}")
                print(
                    f"Booking data: {booking_data.get('booking_id')}, Check-in: {booking_data.get('check_in')}"
                )

                result = send_booking_confirmation(user_email, booking_data)
                print(f"Booking confirmation result: {result}")

                schedule_booking_reminders(
                    booking_data["booking_id"],
                    session["user_id"],
                    booking_data["check_in"],
                    user_email,
                )
                print(f"Scheduled booking reminders for {user_email}")
            except Exception as e:
                print(f"SNS notification error: {e}")
                import traceback

                traceback.print_exc()

            flash("Booking confirmed! Check your email for details.", "success")
            return redirect(
                url_for("booking_success", booking_id=booking_data["booking_id"])
            )

        except Exception as e:
            print(f"Booking error: {e}")
            flash(f"Booking error: {str(e)}", "danger")
            return redirect(url_for("book_room", room_id=room_id))

    # GET request - show booking form
    try:
        # Get room details
        all_rooms = get_all_rooms()
        room = next((r for r in all_rooms if r.get("room_id") == room_id), None)

        if not room:
            flash("Room not found", "danger")
            return redirect(url_for("rooms"))

        # Convert Decimal price to float for template
        if "price" in room:
            room["price"] = convert_decimal_to_float(room["price"])

        return render_template(
            "book_room.html", room=room, user=session.get("user_info", {})
        )
    except Exception as e:
        print(f"Error loading room: {e}")
        flash(f"Error loading room: {str(e)}", "danger")
        return redirect(url_for("rooms"))


@app.route("/booking/success/<booking_id>")
@login_required
def booking_success(booking_id):
    """Booking confirmation page"""
    try:
        # Get the booking details
        booking = get_booking(booking_id)
        if not booking:
            flash("Booking not found", "danger")
            return redirect(url_for("my_bookings"))

        # Get room details
        room = None
        if booking.get("room_id"):
            all_rooms = get_all_rooms()
            room = next(
                (r for r in all_rooms if r.get("room_id") == booking.get("room_id")),
                None,
            )

        return render_template(
            "success.html",
            booking_id=booking_id,
            booking=booking,
            room=room,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading booking success: {e}")
        return render_template(
            "success.html",
            booking_id=booking_id,
            booking=None,
            room=None,
            user=session.get("user_info", {}),
        )


@app.route("/my-bookings")
@login_required
def my_bookings():
    """View user's bookings"""
    try:
        user_id = session.get("user_id")
        if not user_id:
            return redirect(url_for("login"))

        bookings = get_user_bookings(user_id)
        return render_template(
            "bookings.html", bookings=bookings, user=session.get("user_info", {})
        )
    except Exception as e:
        print(f"Error loading bookings: {e}")
        flash(f"Error loading bookings: {str(e)}", "danger")
        return render_template(
            "bookings.html", bookings=[], user=session.get("user_info", {})
        )


@app.route("/room/create", methods=["GET", "POST"])
@login_required
@role_required("admin")
def create_room():
    """Create new room (admin only)"""
    if request.method == "POST":
        try:
            print("=== ROOM CREATION DEBUG ===")
            print(f"User ID: {session.get('user_id')}")
            print(f"User Role: {session.get('role')}")
            print(f"Form Data: {request.form}")
            print(f"Files: {request.files}")

            # Get form data
            name = request.form.get("name", "").strip()
            location = request.form.get("location", "").strip()
            price_str = request.form.get("price", "0").strip()
            capacity_str = request.form.get("capacity", "1").strip()
            amenities_str = request.form.get("amenities", "").strip()
            room_type = request.form.get("room_type", "").strip().lower()
            image_url = request.form.get("image_url", "").strip()
            virtual_tour_360_url = request.form.get("virtual_tour_360_url", "").strip()
            matterport_url = request.form.get("matterport_url", "").strip()

            print(f"Name: {name}")
            print(f"Location: {location}")
            print(f"Price: {price_str}")
            print(f"Capacity: {capacity_str}")
            print(f"Amenities: {amenities_str}")

            # Validate required fields first
            if not name:
                flash("Room name is required", "danger")
                return redirect(url_for("create_room"))

            if not location:
                flash("Room location is required", "danger")
                return redirect(url_for("create_room"))

            # Convert and validate price
            try:
                price = float(price_str)
                if price <= 0:
                    flash("Price must be greater than 0 (in INR)", "danger")
                    return redirect(url_for("create_room"))
            except ValueError:
                flash("Price must be a valid number (in Indian Rupees)", "danger")
                return redirect(url_for("create_room"))

            # Convert and validate capacity
            try:
                capacity = int(capacity_str)
                if capacity < 1:
                    flash("Capacity must be at least 1", "danger")
                    return redirect(url_for("create_room"))
            except ValueError:
                flash("Capacity must be a valid number", "danger")
                return redirect(url_for("create_room"))

            # Process amenities
            amenities = (
                [a.strip() for a in amenities_str.split(",") if a.strip()]
                if amenities_str
                else []
            )

            # Process image (URL wins over upload)
            image_value = None

            if image_url:
                image_value = image_url
            else:
                image_filename = None
                if "room_image" in request.files:
                    file = request.files["room_image"]
                    if file and file.filename:
                        if not allowed_file(file.filename):
                            flash(
                                "Invalid image format. Allowed: PNG, JPG, JPEG, GIF, WEBP",
                                "danger",
                            )
                            return redirect(url_for("create_room"))

                        image_filename = save_room_image(file)
                        if not image_filename:
                            flash("Error saving image. Please try again.", "danger")
                            return redirect(url_for("create_room"))
                image_value = image_filename if image_filename else "default-room.jpg"

            # 360 tour default if not provided
            if not virtual_tour_360_url:
                virtual_tour_360_url = get_360_tour_image(
                    room_type or "default", hash(name + location) % 10
                )

            # Normalize empty matterport to None
            if not matterport_url:
                matterport_url = None

            room_data = {
                "room_id": str(uuid.uuid4()),
                "name": name,
                "location": location,
                "price": price,
                "capacity": capacity,
                "amenities": amenities,
                "availability": "available",
                "image": image_value,
                "room_type": room_type if room_type else "default",
                "virtual_tour_360_url": virtual_tour_360_url,
                "matterport_url": matterport_url,
                "created_by": session.get("user_id", ""),
                "created_at": datetime.now().isoformat(),
            }

            print(f"Room Data: {room_data}")

            # Add room to database
            response = add_room(room_data)
            print(f"Database Response: {response}")

            flash(f'Room "{name}" created successfully!', "success")
            return redirect(url_for("admin_dashboard"))

        except Exception as e:
            print(f"Error creating room: {str(e)}")
            import traceback

            traceback.print_exc()
            flash(f"Error creating room: {str(e)}", "danger")
            return redirect(url_for("create_room"))

    # GET request - show form
    return render_template("create_room.html", user=session.get("user_info", {}))


@app.route("/room/<room_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("admin")
def edit_room(room_id):
    """Edit room details (admin only)"""
    try:
        all_rooms = get_all_rooms()
        room = next((r for r in all_rooms if r.get("room_id") == room_id), None)

        if not room:
            flash("Room not found", "danger")
            return redirect(url_for("admin_dashboard"))

        if request.method == "POST":
            try:
                # Get form data
                name = request.form.get("name", "").strip()
                location = request.form.get("location", "").strip()
                price_str = request.form.get("price", "0").strip()
                capacity_str = request.form.get("capacity", "1").strip()
                amenities_str = request.form.get("amenities", "").strip()
                availability = request.form.get("availability", "available")
                image_url = request.form.get("image_url", "").strip()
                virtual_tour_360_url = request.form.get(
                    "virtual_tour_360_url", ""
                ).strip()
                matterport_url = request.form.get("matterport_url", "").strip()

                # Validate required fields
                if not name:
                    flash("Room name is required", "danger")
                    return redirect(url_for("edit_room", room_id=room_id))

                if not location:
                    flash("Room location is required", "danger")
                    return redirect(url_for("edit_room", room_id=room_id))

                # Convert and validate price
                try:
                    price = float(price_str)
                    if price <= 0:
                        flash("Price must be greater than 0 (in INR)", "danger")
                        return redirect(url_for("edit_room", room_id=room_id))
                except ValueError:
                    flash("Price must be a valid number", "danger")
                    return redirect(url_for("edit_room", room_id=room_id))

                # Convert and validate capacity
                try:
                    capacity = int(capacity_str)
                    if capacity < 1:
                        flash("Capacity must be at least 1", "danger")
                        return redirect(url_for("edit_room", room_id=room_id))
                except ValueError:
                    flash("Capacity must be a valid number", "danger")
                    return redirect(url_for("edit_room", room_id=room_id))

                # Process amenities
                amenities = (
                    [a.strip() for a in amenities_str.split(",") if a.strip()]
                    if amenities_str
                    else []
                )

                # Process image (URL wins over upload; empty keeps existing)
                image_value = room.get("image", "default-room.jpg")
                if image_url:
                    image_value = image_url
                else:
                    if "room_image" in request.files:
                        file = request.files["room_image"]
                        if file and file.filename:
                            if not allowed_file(file.filename):
                                flash(
                                    "Invalid image format. Allowed: PNG, JPG, JPEG, GIF, WEBP",
                                    "danger",
                                )
                                return redirect(url_for("edit_room", room_id=room_id))

                            new_image = save_room_image(file)
                            if new_image:
                                image_value = new_image

                # 360 tour: if empty keep existing; if missing entirely, set default
                if virtual_tour_360_url:
                    tour_value = virtual_tour_360_url
                else:
                    tour_value = room.get("virtual_tour_360_url") or get_360_tour_image(
                        room.get("room_type", "default"), hash(room_id) % 10
                    )

                # Matterport: if provided set; if empty keep existing unless it's blank string
                if matterport_url:
                    matterport_value = matterport_url
                elif "matterport_url" in request.form:
                    # field present but empty -> clear
                    matterport_value = None
                else:
                    matterport_value = room.get("matterport_url")

                # Update room in database
                updated_room = {
                    "room_id": room_id,
                    "name": name,
                    "location": location,
                    "price": price,
                    "capacity": capacity,
                    "amenities": amenities,
                    "availability": availability,
                    "image": image_value,
                    "virtual_tour_360_url": tour_value,
                    "matterport_url": matterport_value,
                    "updated_by": session.get("user_id", ""),
                    "updated_at": datetime.now().isoformat(),
                }

                # Persist update
                update_room(updated_room)

                flash(f'Room "{name}" updated successfully!', "success")
                return redirect(url_for("admin_dashboard"))

            except Exception as e:
                print(f"Error updating room: {str(e)}")
                import traceback

                traceback.print_exc()
                flash(f"Error updating room: {str(e)}", "danger")
                return redirect(url_for("edit_room", room_id=room_id))

        # GET request - show edit form
        return render_template(
            "edit_room.html", room=room, user=session.get("user_info", {})
        )

    except Exception as e:
        print(f"Error loading room: {e}")
        flash(f"Error loading room: {str(e)}", "danger")
        return redirect(url_for("admin_dashboard"))


@app.route("/room/update/<room_id>", methods=["GET", "POST"])
@login_required
@role_required("staff")
def update_availability(room_id):
    """Update room status (staff only)"""
    if request.method == "POST":
        try:
            availability = request.form.get("availability", "available")

            # Validate status
            valid_statuses = ["available", "occupied", "cleaning", "maintenance"]
            if availability not in valid_statuses:
                flash("Invalid room status", "danger")
                return redirect(url_for("staff_dashboard"))

            update_room_availability(room_id, availability)

            status_labels = {
                "available": "✓ Available",
                "occupied": "👤 Occupied",
                "cleaning": "🧹 Cleaning",
                "maintenance": "🔧 Maintenance",
            }

            flash(
                f"Room status updated to {status_labels.get(availability, availability)}!",
                "success",
            )
            return redirect(url_for("staff_dashboard"))
        except Exception as e:
            flash(f"Error updating room: {str(e)}", "danger")
            return redirect(url_for("staff_dashboard"))

    # GET request - show form
    try:
        all_rooms = get_all_rooms()
        room = next((r for r in all_rooms if r.get("room_id") == room_id), None)

        if not room:
            flash("Room not found", "danger")
            return redirect(url_for("staff_dashboard"))

        return render_template(
            "update_room_status.html", room=room, user=session.get("user_info", {})
        )
    except Exception as e:
        flash(f"Error loading room: {str(e)}", "danger")
        return redirect(url_for("staff_dashboard"))


@app.route("/placeholder-room-image")
def placeholder_room_image():
    """Serve placeholder image for missing room images"""
    from flask import send_file
    from io import BytesIO

    # Create SVG placeholder
    svg = """<svg xmlns="http://www.w3.org/2000/svg" width="400" height="250">
        <defs>
            <linearGradient id="grad" x1="0%" y1="0%" x2="100%" y2="100%">
                <stop offset="0%" style="stop-color:#667eea;stop-opacity:1" />
                <stop offset="100%" style="stop-color:#764ba2;stop-opacity:1" />
            </linearGradient>
        </defs>
        <rect fill="url(#grad)" width="400" height="250"/>
        <text x="200" y="80" font-size="60" fill="white" text-anchor="middle" font-family="Arial">🏨</text>
        <text x="200" y="140" font-size="24" fill="white" text-anchor="middle" font-family="Arial">Blissful Abodes</text>
        <text x="200" y="165" font-size="14" fill="rgba(255,255,255,0.7)" text-anchor="middle" font-family="Arial">Room Image</text>
    </svg>"""

    return send_file(
        BytesIO(svg.encode("utf-8")), mimetype="image/svg+xml", as_attachment=False
    )


@app.route("/reviews")
def reviews():
    """View all reviews"""
    try:
        all_reviews = get_all_reviews()
        return render_template(
            "reviews.html", reviews=all_reviews, user=session.get("user_info", {})
        )
    except Exception as e:
        print(f"Error loading reviews: {e}")
        flash(f"Error loading reviews: {str(e)}", "danger")
        return render_template(
            "reviews.html", reviews=[], user=session.get("user_info", {})
        )


@app.route("/room/<room_id>/reviews")
def room_reviews(room_id):
    """View reviews for a specific room"""
    try:
        room = next((r for r in get_all_rooms() if r.get("room_id") == room_id), None)
        if not room:
            flash("Room not found", "danger")
            return redirect(url_for("rooms"))

        reviews = get_room_reviews(room_id)
        rating = get_room_rating(room_id)

        return render_template(
            "room_reviews.html",
            room=room,
            reviews=reviews,
            rating=rating,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading room reviews: {e}")
        flash(f"Error loading reviews: {str(e)}", "danger")
        return redirect(url_for("rooms"))


@app.route("/api/review", methods=["POST"])
@login_required
def api_add_review():
    """API endpoint to add a review"""
    try:
        review_data = {
            "review_id": str(uuid.uuid4()),
            "user_id": session["user_id"],
            "room_id": request.form.get("room_id", ""),
            "rating": int(request.form.get("rating", 0)),
            "title": request.form.get("review_title", ""),
            "text": request.form.get("review_text", ""),
            "helpful_count": 0,
            "created_at": datetime.now().isoformat(),
            "guest_name": session["user_info"].get("name", "Anonymous"),
            "guest_email": session["user_info"].get("email", ""),
        }

        # Validate
        if not all(
            [
                review_data["room_id"],
                review_data["rating"],
                review_data["title"],
                review_data["text"],
            ]
        ):
            return jsonify({"success": False, "error": "All fields required"})

        if review_data["rating"] < 1 or review_data["rating"] > 5:
            return jsonify(
                {"success": False, "error": "Rating must be between 1 and 5"}
            )

        add_review(review_data)
        return jsonify({"success": True, "message": "Review added successfully"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# API endpoints for AJAX requests
@app.route("/api/rooms/available")
def api_available_rooms():
    """API endpoint for available rooms"""
    try:
        location = request.args.get("location", "")
        rooms = get_rooms(location)
        return jsonify({"success": True, "rooms": rooms})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/analytics")
@login_required
@role_required("admin")
def api_analytics():
    """API endpoint for analytics data"""
    try:
        analytics = get_analytics()
        return jsonify({"success": True, "data": analytics})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/users", methods=["GET", "POST"])
@login_required
@role_required("admin")
def api_users():
    """Get all users or create new user"""
    if request.method == "GET":
        try:
            users = get_all_users()

            formatted_users = []
            for user in users:
                formatted_users.append(
                    {
                        "id": user.get("user_id", ""),
                        "user_id": user.get("user_id", ""),
                        "name": user.get("name", "N/A"),
                        "email": user.get("email", ""),
                        "role": user.get("role", "guest"),
                        "age": user.get("age", 0),
                        "phone": user.get("phone", ""),
                        "joined": (
                            user.get("created_at", "")[:10]
                            if user.get("created_at")
                            else "Unknown"
                        ),
                    }
                )

            print(f"✓ Returning {len(formatted_users)} users")
            return jsonify({"success": True, "users": formatted_users})
        except Exception as e:
            print(f"Error getting users: {e}")
            import traceback

            traceback.print_exc()
            return jsonify({"success": False, "error": str(e)})

    elif request.method == "POST":
        try:
            user_data = {
                "user_id": str(uuid.uuid4()),
                "name": request.form.get("name", "").strip(),
                "email": request.form.get("email", "").strip().lower(),
                "password": hash_password(request.form.get("password", "")),
                "age": int(request.form.get("age", 0)),
                "role": request.form.get("role", "guest"),
                "phone": request.form.get("phone", ""),
                "created_at": datetime.now().isoformat(),
            }

            # Validate
            if not all([user_data["name"], user_data["email"], user_data["password"]]):
                return jsonify({"success": False, "error": "All fields required"})

            if user_data["age"] < 18:
                return jsonify({"success": False, "error": "Must be 18+"})

            existing = get_user_by_email(user_data["email"])
            if existing:
                return jsonify({"success": False, "error": "Email already exists"})

            add_user(user_data)
            print(f"✓ Created user: {user_data['email']}")
            return jsonify({"success": True, "message": "User created successfully"})
        except Exception as e:
            print(f"Error creating user: {e}")
            return jsonify({"success": False, "error": str(e)})


@app.route("/api/user/<user_id>", methods=["GET", "PUT", "DELETE"])
@login_required
@role_required("admin")
def api_user_detail(user_id):
    """Get/update/delete a single user (used by admin dashboard)."""
    try:
        if request.method == "GET":
            user = get_user(user_id)
            if not user:
                return jsonify({"success": False, "error": "User not found"}), 404
            return jsonify({"success": True, "user": user})

        if request.method == "PUT":
            payload = request.get_json(silent=True) or {}

            update_data = {}
            for key in ["name", "email", "role", "age", "phone", "branch_id"]:
                if key in payload:
                    update_data[key] = payload.get(key)

            # Normalize email
            if "email" in update_data and isinstance(update_data["email"], str):
                update_data["email"] = update_data["email"].strip().lower()

            # Normalize age
            if "age" in update_data:
                try:
                    update_data["age"] = int(update_data["age"])
                except (TypeError, ValueError):
                    return (
                        jsonify({"success": False, "error": "Age must be a number"}),
                        400,
                    )

            # Optional password change
            new_password = payload.get("password")
            if isinstance(new_password, str) and new_password.strip():
                update_data["password"] = hash_password(new_password.strip())

            update_data["updated_at"] = datetime.now().isoformat()

            update_user(user_id, update_data)
            return jsonify({"success": True, "message": "User updated successfully"})

        # DELETE
        delete_user(user_id)
        return jsonify({"success": True, "message": "User deleted successfully"})

    except Exception as e:
        print(f"Error in user detail API: {e}")
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/rooms", methods=["GET"])
@login_required
@role_required("admin")
def api_rooms():
    """Get all rooms"""
    try:
        rooms = get_all_rooms()

        formatted_rooms = []
        for room in rooms:
            formatted_rooms.append(
                {
                    "id": room.get("room_id", ""),
                    "room_id": room.get("room_id", ""),
                    "name": room.get("name", "N/A"),
                    "location": room.get("location", ""),
                    "price": float(room.get("price", 0)),
                    "capacity": int(room.get("capacity", 1)),
                    "status": room.get("availability", "available"),
                    "availability": room.get("availability", "available"),
                }
            )

        print(f"✓ Returning {len(formatted_rooms)} rooms")
        return jsonify({"success": True, "rooms": formatted_rooms})
    except Exception as e:
        print(f"Error getting rooms: {e}")
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/room/<room_id>", methods=["DELETE"])
@login_required
@role_required("admin")
def api_room_detail(room_id):
    """Delete a single room (used by admin dashboard)"""
    try:
        delete_room(room_id)
        return jsonify({"success": True, "message": "Room deleted successfully"})
    except Exception as e:
        print(f"Error deleting room: {e}")
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/bookings", methods=["GET"])
@login_required
@role_required("admin")
def api_bookings_list():
    """Get all bookings"""
    try:
        bookings = get_all_bookings()
        formatted_bookings = []
        for booking in bookings:
            formatted_bookings.append(
                {
                    "booking_id": booking.get("booking_id", ""),
                    "room_id": booking.get("room_id", ""),
                    "user_id": booking.get("user_id", ""),
                    "guest_name": booking.get("guest_name", "N/A"),
                    "guest_email": booking.get("guest_email", ""),
                    "check_in": booking.get("check_in", ""),
                    "check_out": booking.get("check_out", ""),
                    "status": booking.get("status", "pending"),
                    "created_at": booking.get("created_at", ""),
                }
            )
        return jsonify({"success": True, "bookings": formatted_bookings})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/booking/<booking_id>", methods=["DELETE", "PUT"])
@login_required
@role_required("admin")
def api_manage_booking(booking_id):
    """Cancel or update booking status"""
    if request.method == "DELETE":
        try:
            # Find booking to get associated room_id
            booking = get_booking(booking_id)
            if not booking:
                return jsonify({"success": False, "error": "Booking not found"}), 404

            room_id = booking.get("room_id")

            # Mark booking as cancelled (do not hard-delete so history is preserved)
            update_booking_status(booking_id, "cancelled")

            # Free the room if we know which one it is
            if room_id:
                update_room_availability(room_id, "available")

            return jsonify({"success": True, "message": "Booking cancelled"})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)})

    elif request.method == "PUT":
        try:
            payload = request.get_json(silent=True) or {}
            new_status = payload.get("status", "confirmed")

            # Update booking status fields
            update_booking_status(booking_id, new_status)

            # Optionally adjust room availability based on status
            booking = get_booking(booking_id)
            room_id = booking.get("room_id") if booking else None
            if room_id:
                if new_status in ["confirmed", "paid"]:
                    update_room_availability(room_id, "unavailable")
                elif new_status in ["cancelled"]:
                    update_room_availability(room_id, "available")

            return jsonify({"success": True, "message": "Booking updated"})
        except Exception as e:
            return jsonify({"success": False, "error": str(e)})


# ============================================================================
# CHATBOT ROUTES
# ============================================================================


@app.route("/chatbot")
@login_required
def chatbot():
    """Chatbot interface for users"""
    user_id = session.get("user_id")
    user = session.get("user_info", {})
    messages = get_user_chat_messages(user_id)
    unread_count = get_unread_messages_count(user_id)

    # Mark messages as read when user opens chat
    mark_messages_as_read(user_id)

    # Get context from referrer or query parameter
    context = request.args.get("context", "general")
    referrer = request.referrer or ""

    # Determine context from referrer URL
    if "rooms" in referrer or context == "rooms":
        page_context = "rooms"
    elif "booking" in referrer or context == "bookings":
        page_context = "bookings"
    elif "review" in referrer or context == "reviews":
        page_context = "reviews"
    elif "dashboard" in referrer or context == "dashboard":
        page_context = "dashboard"
    else:
        page_context = "general"

    return render_template(
        "chatbot.html",
        user=user,
        messages=messages,
        unread_count=unread_count,
        page_context=page_context,
    )


@app.route("/api/chatbot/send", methods=["POST"])
@login_required
def api_chatbot_send():
    """Send a message in chatbot"""
    try:
        user_id = session.get("user_id")
        user = session.get("user_info", {})
        data = request.get_json()
        message_text = data.get("message", "").strip()

        if not message_text:
            return jsonify({"success": False, "error": "Message cannot be empty"})

        # Add user message
        user_message = add_chat_message(user_id, message_text, sender="user")

        # Get context from request
        context = data.get("context", "general")

        # Process message and generate bot response
        bot_response = process_chat_message(user, message_text, context)

        # Add bot message
        bot_message = add_chat_message(
            user_id,
            bot_response["message"],
            sender="bot",
            message_type=bot_response.get("type", "text"),
        )

        return jsonify(
            {
                "success": True,
                "user_message": user_message,
                "bot_message": bot_message,
                "request_created": bot_response.get("request_created", False),
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chatbot/messages")
@login_required
def api_chatbot_messages():
    """Get chat messages for current user"""
    try:
        user_id = session.get("user_id")
        messages = get_user_chat_messages(user_id)
        return jsonify({"success": True, "messages": messages})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/chatbot/request", methods=["POST"])
@login_required
def api_chatbot_request():
    """Create a chat request (booking, review, report, extra service)"""
    try:
        user_id = session.get("user_id")
        user = session.get("user_info", {})
        data = request.get_json()

        request_type = data.get(
            "type"
        )  # 'booking', 'review', 'report', 'extra_service'
        details = data.get("details", {})
        branch_id = data.get("branch_id")

        if not request_type:
            return jsonify({"success": False, "error": "Request type is required"})

        # Create the request
        chat_request = add_chat_request(user_id, request_type, details, branch_id)

        if not chat_request:
            return jsonify({"success": False, "error": "Failed to create request"})

        # Send notification to admin
        notify_admin_of_request(chat_request, user)

        # If extra service, notify staff
        if request_type == "extra_service":
            notify_staff_of_service_request(chat_request, user)

        return jsonify(
            {
                "success": True,
                "request": chat_request,
                "message": "Your request has been submitted successfully",
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/admin/chatbot")
@login_required
@role_required(["admin", "super_admin", "branch_manager"])
def admin_chatbot():
    """Admin chatbot management dashboard"""
    user = session.get("user_info", {})

    # Get all requests
    all_requests = get_chat_requests()
    pending_requests = get_chat_requests(status="pending")
    processing_requests = get_chat_requests(status="processing")

    # Get request counts by type
    booking_requests = [r for r in all_requests if r["request_type"] == "booking"]
    review_requests = [r for r in all_requests if r["request_type"] == "review"]
    report_requests = [r for r in all_requests if r["request_type"] == "report"]
    extra_service_requests = [
        r for r in all_requests if r["request_type"] == "extra_service"
    ]

    return render_template(
        "admin_chatbot.html",
        user=user,
        all_requests=all_requests,
        pending_requests=pending_requests,
        processing_requests=processing_requests,
        booking_requests=booking_requests,
        review_requests=review_requests,
        report_requests=report_requests,
        extra_service_requests=extra_service_requests,
    )


@app.route("/api/admin/chatbot/request/<request_id>", methods=["GET", "PUT", "DELETE"])
@login_required
@role_required(["admin", "super_admin", "branch_manager"])
def api_admin_chatbot_request(request_id):
    """Manage a specific chat request"""
    try:
        if request.method == "GET":
            chat_request = get_chat_request(request_id)
            if not chat_request:
                return jsonify({"success": False, "error": "Request not found"})
            return jsonify({"success": True, "request": chat_request})

        elif request.method == "PUT":
            data = request.get_json()
            status = data.get("status")
            admin_notes = data.get("admin_notes")
            staff_assigned = data.get("staff_assigned")

            updated_request = update_chat_request(
                request_id, status, admin_notes, staff_assigned
            )

            if not updated_request:
                return jsonify({"success": False, "error": "Failed to update request"})

            # Auto-process based on status
            if status == "completed":
                autoprocess_completed_request(updated_request)

            return jsonify({"success": True, "request": updated_request})

        elif request.method == "DELETE":
            success = delete_chat_request(request_id)
            return jsonify({"success": success})

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/admin/chatbot/requests")
@login_required
@role_required(["admin", "super_admin", "branch_manager"])
def api_admin_chatbot_requests():
    """Get all chat requests with filters"""
    try:
        status = request.args.get("status")
        request_type = request.args.get("type")

        requests = get_chat_requests(status=status, request_type=request_type)
        return jsonify({"success": True, "requests": requests})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/staff/chatbot")
@login_required
@role_required(["staff", "branch_staff", "branch_manager", "admin", "super_admin"])
def staff_chatbot():
    """Staff chatbot dashboard for extra services"""
    user = session.get("user_info", {})

    # Get extra service requests
    all_service_requests = get_chat_requests(request_type="extra_service")

    # Filter by branch if branch staff
    if user["role"] in ["branch_staff", "branch_manager"]:
        branch_id = user.get("branch_id")
        if branch_id:
            all_service_requests = [
                r for r in all_service_requests if r.get("branch_id") == branch_id
            ]

    pending_services = [r for r in all_service_requests if r["status"] == "pending"]
    processing_services = [
        r for r in all_service_requests if r["status"] == "processing"
    ]
    completed_services = [r for r in all_service_requests if r["status"] == "completed"]

    return render_template(
        "staff_chatbot.html",
        user=user,
        all_service_requests=all_service_requests,
        pending_services=pending_services,
        processing_services=processing_services,
        completed_services=completed_services,
    )


# ============================================================================
# CHATBOT HELPER FUNCTIONS
# ============================================================================


def process_chat_message(user, message, context="general"):
    """Process user message and generate bot response based on page context"""
    message_lower = message.lower()

    # helper for checking multiple keywords
    def contains_any(keywords):
        return any(word in message_lower for word in keywords)

    # --- 1. GREETINGS ---
    if contains_any(
        ["hi", "hello", "hey", "greetings", "morning", "evening", "afternoon"]
    ):
        context_greeting = ""
        if context == "rooms":
            context_greeting = " I see you're looking for a room. Need recommendations?"
        elif context == "bookings":
            context_greeting = " checking your reservations?"
        elif context == "reviews":
            context_greeting = " looking to share your feedback?"

        return {
            "message": f"Hello {user.get('name', 'Guest')}! Welcome to Blissful Abodes Assistant. 👋{context_greeting}\n\nI can help you with:\n• 🏨 Booking a Room\n• 📅 Managing Reservations\n• ℹ️ Hotel Amenities & Policies\n• ⭐ Reviews & Feedback\n\nHow can I help you today?",
            "type": "text",
        }

    # --- 2. AMENITIES & FACILITIES ---
    if contains_any(
        [
            "pool",
            "swim",
            "gym",
            "fitness",
            "wifi",
            "internet",
            "parking",
            "breakfast",
            "food",
            "restaurant",
            "spa",
        ]
    ):
        response = "Here's what we offer regarding your query: 🏨\n\n"
        if contains_any(["pool", "swim"]):
            response += "• 🏊‍♂️ **Swimming Pool**: We have a temperature-controlled infinity pool open from 6 AM to 10 PM.\n"
        if contains_any(["gym", "fitness", "workout"]):
            response += "• 🏋️ **Gym**: Our state-of-the-art fitness center is open 24/7 for guests.\n"
        if contains_any(["wifi", "internet"]):
            response += "• 📶 **Wi-Fi**: High-speed complimentary Wi-Fi is available throughout the property.\n"
        if contains_any(["parking", "car"]):
            response += (
                "• 🚗 **Parking**: We offer free secure valet parking for all guests.\n"
            )
        if contains_any(["breakfast", "food", "restaurant", "dining"]):
            response += "• 🍽️ **Dining**: We have an in-house multi-cuisine restaurant and 24/7 room service. Breakfast is complimentary with most plans.\n"
        if contains_any(["spa", "massage"]):
            response += "• 💆 **Spa**: Our luxury spa offers various treatments from 9 AM to 8 PM.\n"

        return {"message": response, "type": "text"}

    # --- 3. POLICIES (Check-in, Cancellation, etc.) ---
    if contains_any(
        [
            "check in",
            "check-in",
            "check out",
            "check-out",
            "time",
            "policy",
            "rules",
            "cancel",
            "refund",
            "pet",
            "dog",
            "cat",
        ]
    ):
        response = "Here are our key policies: 📋\n\n"
        if contains_any(["check in", "check-in", "check out", "check-out", "time"]):
            response += "• 🕒 **Check-in/Out**: Check-in time is 2:00 PM and Check-out time is 11:00 AM.\n"
        if contains_any(["cancel", "refund"]):
            response += "• 🔄 **Cancellation**: Free cancellation up to 24 hours before check-in. Late cancellations may incur a one-night charge.\n"
        if contains_any(["pet", "dog", "cat", "animal"]):
            response += "• 🐾 **Pets**: We are a pet-friendly hotel! A small cleaning fee may apply.\n"
        if contains_any(["smoke", "smoking"]):
            response += "• 🚭 **Smoking**: All rooms are non-smoking. Designated smoking areas are available.\n"
        if contains_any(["children", "kids", "baby"]):
            response += (
                "• 👶 **Children**: Kids under 5 stay free using existing bedding.\n"
            )

        return {"message": response, "type": "text"}

    # --- 4. LOCATION & CONTACT ---
    if contains_any(
        [
            "where",
            "location",
            "address",
            "map",
            "contact",
            "phone",
            "email",
            "call",
            "reach",
        ]
    ):
        return {
            "message": "📍 **Location & Contact**:\n\nWe are located at:\n123 Cloud Avenue, Tech Park District\nAWS Region, Digital India\n\n📞 **Phone**: +91 98765 43210\n📧 **Email**: support@blissfulabodes.com\n\nWe are centrally located near major attractions and the airport!",
            "type": "text",
        }

    # --- 5. BOOKING INTENT ---
    if contains_any(
        [
            "book",
            "reservation",
            "reserve",
            "room",
            "stay",
            "price",
            "cost",
            "rate",
            "availability",
            "available",
        ]
    ):
        if context == "rooms":
            return {
                "message": "You're in the right place! 🏨\n\nYou can:\n1. Select your dates in the search bar above\n2. Filter by room type\n3. Click 'Book Now' on your preferred room\n\nNeed help checking availability for specific dates?",
                "type": "booking",
            }
        return {
            "message": "I can help you book a stay! 🏨\n\nPlease provide:\n1. Check-in Date\n2. Check-out Date\n3. Number of Guests\n\nOr simply click the button below to start:",
            "type": "booking",
        }

    # --- 6. EXISTING BOOKING MANAGEMENT ---
    if contains_any(
        ["my booking", "cancel booking", "modify", "change booking", "status"]
    ):
        return {
            "message": "You can manage your bookings in the 'My Bookings' section. 📅\n\nThere you can:\n• View booking status\n• Download receipts\n• Cancel or modify reservations\n\nWould you like me to take you there?",
            "type": "report",  # Keeping generic type for specific actions
        }

    # --- 7. REVIEWS & FEEDBACK ---
    if contains_any(
        ["review", "rating", "feedback", "star", "complain", "issue", "report"]
    ):
        if contains_any(
            ["complain", "issue", "report", "problem", "mad", "angry", "bad"]
        ):
            return {
                "message": "I'm very sorry to hear you're facing an issue. 😟\n\nPlease describe the problem in detail using the 'Report Issue' button so our team can resolve it immediately.",
                "type": "report",
            }
        return {
            "message": "We value your feedback! ⭐\n\nYou can rate your stay and write a review after your checkout. Use the 'Submit Review' button to share your experience!",
            "type": "review",
        }

    # --- 8. EXTRA SERVICES ---
    if contains_any(
        ["service", "taxi", "cab", "tour", "clean", "laundry", "doctor", "help"]
    ):
        return {
            "message": "We offer a range of premium services: 🛎️\n\n• Room Service & Dining\n• Spa & Wellness\n• Laundry & Dry Cleaning\n• Airport Transfers & Taxi\n• City Tours\n• Doctor on Call\n\nClick 'Request Service' to book any of these.",
            "type": "extra_service",
        }

    # --- 9. FALLBACK / DEFAULT ---
    # Try to keep the conversation going based on context
    fallback_msg = "I'm not sure I understood that correctly. 🤔\n\nI can help with:"
    options = "\n• 🏨 Booking a Room\n• ℹ️ Hotel Information & Amenities\n• 📅 Managing Reservations\n• ⭐ Reviews & Feedback"

    if context == "rooms":
        fallback_msg = "I'm not sure, but I can help you find a room! 🏨"
        options = "\n• Check Availability\n• View Room Types\n• Pricing & Offers"
    elif context == "bookings":
        fallback_msg = "I'm here to help with your bookings. 📅"
        options = "\n• View Confirmation\n• Cancel/Modify\n• Download Receipt"

    return {
        "message": f"{fallback_msg}{options}\n\nPlease try asking in a different way or use the quick buttons.",
        "type": "text",
    }


def notify_admin_of_request(chat_request, user):
    """Notify admin of new chat request"""
    try:
        request_type = chat_request["request_type"].replace("_", " ").title()
        subject = f"New {request_type} Request from {user.get('name', 'User')}"
        message = f"""
New chat request received:

Type: {request_type}
User: {user.get('name', 'Unknown')} ({user.get('email', 'Unknown')})
Request ID: {chat_request['request_id']}
Created: {chat_request['created_at']}

Details:
{json.dumps(chat_request.get('details', {}), indent=2)}

Please review and process this request in the admin dashboard.
        """
        send_notification(subject, message)
    except Exception as e:
        print(f"Error notifying admin: {e}")


def notify_staff_of_service_request(chat_request, user):
    """Notify staff of new extra service request"""
    try:
        subject = f"New Service Request from {user.get('name', 'User')}"
        message = f"""
New extra service request:

User: {user.get('name', 'Unknown')} ({user.get('email', 'Unknown')})
Request ID: {chat_request['request_id']}
Branch: {chat_request.get('branch_id', 'Not specified')}
Created: {chat_request['created_at']}

Service Details:
{json.dumps(chat_request.get('details', {}), indent=2)}

Please process this service request as soon as possible.
        """
        send_notification(subject, message)

        # Also add a message to the chat
        user_id = chat_request.get("user_id")
        if user_id:
            add_chat_message(
                user_id,
                "Your service request has been forwarded to our staff. They will contact you shortly!",
                sender="bot",
            )
    except Exception as e:
        print(f"Error notifying staff: {e}")


def autoprocess_completed_request(chat_request):
    """Auto-process completed requests"""
    try:
        request_type = chat_request["request_type"]
        details = chat_request.get("details", {})
        user_id = chat_request["user_id"]

        # If booking request is completed, create the booking
        if request_type == "booking" and details.get("room_id"):
            booking_data = {
                "user_id": user_id,
                "room_id": details["room_id"],
                "check_in": details.get("check_in"),
                "check_out": details.get("check_out"),
                "guests": details.get("guests", 1),
                "total_price": details.get("total_price", 0),
                "status": "confirmed",
            }
            add_booking(booking_data)
            add_chat_message(
                user_id,
                f"Your booking has been confirmed! Booking ID: {booking_data.get('booking_id', 'N/A')}",
                sender="bot",
            )

        # If review request is completed, create the review
        elif request_type == "review" and details.get("room_id"):
            review_data = {
                "user_id": user_id,
                "room_id": details["room_id"],
                "rating": details.get("rating", 5),
                "comment": details.get("comment", ""),
                "date": datetime.now().isoformat(),
            }
            add_review(review_data)
            add_chat_message(
                user_id,
                "Thank you for your review! It has been published.",
                sender="bot",
            )

        # For reports and extra services, just send confirmation
        else:
            add_chat_message(
                user_id,
                f"Your {request_type.replace('_', ' ')} request has been completed!",
                sender="admin",
            )

    except Exception as e:
        print(f"Error auto-processing request: {e}")


# ============================================================================
# REPORT GENERATION
# ============================================================================


@app.route("/api/admin/generate-report")
@login_required
@role_required(["admin", "super_admin"])
def generate_system_report():
    """Generate comprehensive system report as Excel"""
    try:
        # Get all system data
        analytics = get_analytics()
        users = get_all_users()
        rooms = get_all_rooms()
        bookings = get_all_bookings()

        # Calculate revenue metrics
        total_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "paid"
        )
        pending_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "pending"
        )
        avg_booking_value = total_revenue / len(bookings) if len(bookings) > 0 else 0

        # Calculate occupancy
        occupied_rooms = len(
            [r for r in rooms if r.get("availability") == "unavailable"]
        )
        occupancy_rate = (occupied_rooms / len(rooms) * 100) if len(rooms) > 0 else 0

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"blissful_abodes_report_{timestamp}.xlsx"

        if OPENPYXL_AVAILABLE:
            wb = openpyxl.Workbook()

            # Summary Sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"

            title_font = Font(name="Arial", size=14, bold=True, color="1A365D")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="1A365D", end_color="1A365D", fill_type="solid"
            )
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            ws_summary["A1"] = "BLISSFUL ABODES SYSTEM REPORT"
            ws_summary["A1"].font = title_font
            ws_summary["A2"] = (
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            )

            summary_data = [
                ["Metric", "Value"],
                ["Total Users", len(users)],
                ["Total Rooms", len(rooms)],
                ["Total Bookings", len(bookings)],
                ["Occupancy Rate", f"{occupancy_rate:.1f}%"],
                ["Total Revenue", f"₹{total_revenue:,.2f}"],
                ["Pending Revenue", f"₹{pending_revenue:,.2f}"],
                ["Avg Booking Value", f"₹{avg_booking_value:,.2f}"],
            ]

            for r_idx, row in enumerate(summary_data, 4):
                for c_idx, val in enumerate(row, 1):
                    cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = border
                    if r_idx == 4:
                        cell.font = header_font
                        cell.fill = header_fill

            # Users Sheet
            ws_users = wb.create_sheet("Users")
            headers = ["Name", "Email", "Role", "Age"]
            for c_idx, head in enumerate(headers, 1):
                cell = ws_users.cell(row=1, column=c_idx, value=head)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for r_idx, user in enumerate(users, 2):
                ws_users.cell(
                    row=r_idx, column=1, value=user.get("name", "N/A")
                ).border = border
                ws_users.cell(
                    row=r_idx, column=2, value=user.get("email", "N/A")
                ).border = border
                ws_users.cell(
                    row=r_idx, column=3, value=user.get("role", "N/A")
                ).border = border
                ws_users.cell(
                    row=r_idx, column=4, value=user.get("age", "N/A")
                ).border = border

            # Rooms Sheet
            ws_rooms = wb.create_sheet("Rooms")
            headers = ["Room Name", "Location", "Price", "Status"]
            for c_idx, head in enumerate(headers, 1):
                cell = ws_rooms.cell(row=1, column=c_idx, value=head)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for r_idx, room in enumerate(rooms, 2):
                ws_rooms.cell(
                    row=r_idx, column=1, value=room.get("name", "N/A")
                ).border = border
                ws_rooms.cell(
                    row=r_idx, column=2, value=room.get("location", "N/A")
                ).border = border
                ws_rooms.cell(
                    row=r_idx, column=3, value=room.get("price", 0)
                ).border = border
                ws_rooms.cell(
                    row=r_idx, column=4, value=room.get("availability", "N/A")
                ).border = border

            # Bookings Sheet
            ws_bookings = wb.create_sheet("Bookings")
            headers = ["Guest", "Room ID", "Check-in", "Check-out", "Price", "Status"]
            for c_idx, head in enumerate(headers, 1):
                cell = ws_bookings.cell(row=1, column=c_idx, value=head)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            for r_idx, b in enumerate(bookings, 2):
                ws_bookings.cell(
                    row=r_idx, column=1, value=b.get("guest_name", "N/A")
                ).border = border
                ws_bookings.cell(
                    row=r_idx, column=2, value=b.get("room_id", "N/A")
                ).border = border
                ws_bookings.cell(
                    row=r_idx, column=3, value=b.get("check_in", "N/A")
                ).border = border
                ws_bookings.cell(
                    row=r_idx, column=4, value=b.get("check_out", "N/A")
                ).border = border
                ws_bookings.cell(
                    row=r_idx, column=5, value=float(b.get("total_price", 0))
                ).border = border
                ws_bookings.cell(
                    row=r_idx, column=6, value=b.get("status", "N/A")
                ).border = border

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return send_file(
                buffer,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )

        # Fallback to CSV
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(["Blissful Abodes System Report"])
        writer.writerow([f"Generated: {datetime.now()}"])
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Users", len(users)])
        writer.writerow(["Total Rooms", len(rooms)])
        writer.writerow(["Total Bookings", len(bookings)])
        writer.writerow(["Total Revenue", total_revenue])

        output.seek(0)
        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = (
            f"attachment; filename={filename.replace('.xlsx', '.csv')}"
        )
        return response

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/export-users")
@login_required
@role_required(["admin", "super_admin"])
def export_users():
    """Export users data as CSV"""
    try:
        users = get_all_users()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"blissful_abodes_users_{timestamp}.csv"

        # Create CSV
        output = StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow(
            [
                "User ID",
                "Name",
                "Email",
                "Role",
                "Age",
                "Phone",
                "Branch ID",
                "Created At",
            ]
        )

        # Data
        for user in users:
            writer.writerow(
                [
                    user.get("user_id", "N/A"),
                    user.get("name", "N/A"),
                    user.get("email", "N/A"),
                    user.get("role", "N/A"),
                    user.get("age", "N/A"),
                    user.get("phone", "N/A"),
                    user.get("branch_id", "N/A"),
                    user.get("created_at", "N/A"),
                ]
            )

        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        print(f"Error exporting users: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/export-bookings")
@login_required
@role_required(["admin", "super_admin"])
def export_bookings():
    """Export bookings data as CSV with real information"""
    try:
        bookings = get_all_bookings()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"blissful_abodes_bookings_{timestamp}.csv"

        # Create CSV
        output = StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow(
            [
                "Booking ID",
                "Guest Name",
                "Guest Email",
                "Room ID",
                "Branch ID",
                "Check-in",
                "Check-out",
                "Nights",
                "Total Price",
                "Currency",
                "Payment Status",
                "Booking Status",
                "Created At",
            ]
        )

        # Data
        for booking in bookings:
            writer.writerow(
                [
                    booking.get("booking_id", "N/A"),
                    booking.get("guest_name", "N/A"),
                    booking.get("guest_email", "N/A"),
                    booking.get("room_id", "N/A"),
                    booking.get("branch_id", "N/A"),
                    booking.get("check_in", "N/A"),
                    booking.get("check_out", "N/A"),
                    booking.get("nights", 0),
                    f"₹{float(booking.get('total_price', 0)):,.2f}",
                    booking.get("currency", "INR"),
                    booking.get("payment_status", "N/A"),
                    booking.get("status", "N/A"),
                    booking.get("created_at", "N/A"),
                ]
            )

        # Add summary at the end
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        total_bookings = len(bookings)
        total_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "paid"
        )
        pending_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "pending"
        )
        confirmed_bookings = len(
            [b for b in bookings if b.get("status") == "confirmed"]
        )

        writer.writerow(["Total Bookings", total_bookings])
        writer.writerow(["Confirmed Bookings", confirmed_bookings])
        writer.writerow(["Total Revenue (Paid)", f"₹{total_revenue:,.2f}"])
        writer.writerow(["Pending Revenue", f"₹{pending_revenue:,.2f}"])
        writer.writerow(
            [
                "Average Booking Value",
                (
                    f"₹{total_revenue/confirmed_bookings:,.2f}"
                    if confirmed_bookings > 0
                    else "₹0.00"
                ),
            ]
        )

        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        print(f"Error exporting bookings: {e}")
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/export-rooms")
@login_required
@role_required(["admin", "super_admin"])
def export_rooms():
    """Export rooms data as CSV with real information"""
    try:
        rooms = get_all_rooms()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"blissful_abodes_rooms_{timestamp}.csv"

        # Create CSV
        output = StringIO()
        writer = csv.writer(output)

        # Headers
        writer.writerow(
            [
                "Room ID",
                "Room Name",
                "Location",
                "Branch ID",
                "Price",
                "Room Type",
                "Capacity",
                "Availability",
                "Amenities",
                "Description",
            ]
        )

        # Data
        for room in rooms:
            writer.writerow(
                [
                    room.get("room_id", "N/A"),
                    room.get("name", "N/A"),
                    room.get("location", "N/A"),
                    room.get("branch_id", "N/A"),
                    f"₹{float(room.get('price', 0)):,.2f}",
                    room.get("room_type", "N/A"),
                    room.get("capacity", "N/A"),
                    room.get("availability", "N/A"),
                    room.get("amenities", "N/A"),
                    (
                        room.get("description", "N/A")[:100]
                        if room.get("description")
                        else "N/A"
                    ),
                ]
            )

        # Add summary at the end
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        total_rooms = len(rooms)
        available_rooms = len(
            [r for r in rooms if r.get("availability") == "available"]
        )
        occupied_rooms = len(
            [r for r in rooms if r.get("availability") == "unavailable"]
        )
        avg_price = (
            sum(float(r.get("price", 0)) for r in rooms) / total_rooms
            if total_rooms > 0
            else 0
        )
        occupancy_rate = (occupied_rooms / total_rooms * 100) if total_rooms > 0 else 0

        writer.writerow(["Total Rooms", total_rooms])
        writer.writerow(["Available Rooms", available_rooms])
        writer.writerow(["Occupied Rooms", occupied_rooms])
        writer.writerow(["Occupancy Rate", f"{occupancy_rate:.1f}%"])
        writer.writerow(["Average Room Price", f"₹{avg_price:,.2f}"])

        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        print(f"Error exporting rooms: {e}")
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/api/admin/export-revenue")
@login_required
@role_required(["admin", "super_admin"])
def export_revenue():
    """Export detailed revenue report as CSV"""
    try:
        bookings = get_all_bookings()
        branches = get_all_branches()

        # Generate filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"blissful_abodes_revenue_{timestamp}.csv"

        # Create CSV
        output = StringIO()
        writer = csv.writer(output)

        # Overall Summary
        writer.writerow(["BLISSFUL ABODES - REVENUE REPORT"])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([])

        # Calculate overall metrics
        total_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "paid"
        )
        pending_revenue = sum(
            float(b.get("total_price", 0))
            for b in bookings
            if b.get("payment_status") == "pending"
        )
        total_bookings = len(bookings)
        paid_bookings = len([b for b in bookings if b.get("payment_status") == "paid"])

        writer.writerow(["OVERALL METRICS"])
        writer.writerow(["Total Bookings", total_bookings])
        writer.writerow(["Paid Bookings", paid_bookings])
        writer.writerow(["Total Revenue (Paid)", f"₹{total_revenue:,.2f}"])
        writer.writerow(["Pending Revenue", f"₹{pending_revenue:,.2f}"])
        writer.writerow(
            [
                "Average Revenue per Booking",
                (
                    f"₹{total_revenue/paid_bookings:,.2f}"
                    if paid_bookings > 0
                    else "₹0.00"
                ),
            ]
        )
        writer.writerow([])

        # Revenue by Branch
        writer.writerow(["REVENUE BY BRANCH"])
        writer.writerow(
            [
                "Branch Name",
                "Location",
                "Total Bookings",
                "Revenue (Paid)",
                "Pending Revenue",
                "Average Booking Value",
            ]
        )

        for branch in branches:
            branch_id = branch.get("branch_id")
            branch_bookings = [b for b in bookings if b.get("branch_id") == branch_id]
            branch_revenue = sum(
                float(b.get("total_price", 0))
                for b in branch_bookings
                if b.get("payment_status") == "paid"
            )
            branch_pending = sum(
                float(b.get("total_price", 0))
                for b in branch_bookings
                if b.get("payment_status") == "pending"
            )
            branch_paid_count = len(
                [b for b in branch_bookings if b.get("payment_status") == "paid"]
            )
            avg_value = (
                branch_revenue / branch_paid_count if branch_paid_count > 0 else 0
            )

            location = (
                branch.get("location", {}).get("city", "N/A")
                if isinstance(branch.get("location"), dict)
                else "N/A"
            )

            writer.writerow(
                [
                    branch.get("branch_name", "N/A"),
                    location,
                    len(branch_bookings),
                    f"₹{branch_revenue:,.2f}",
                    f"₹{branch_pending:,.2f}",
                    f"₹{avg_value:,.2f}",
                ]
            )

        writer.writerow([])

        # Revenue by Month (last 12 months)
        writer.writerow(["MONTHLY REVENUE TREND"])
        writer.writerow(["Month", "Bookings", "Revenue"])

        # Group bookings by month
        from collections import defaultdict

        monthly_data = defaultdict(lambda: {"count": 0, "revenue": 0})

        for booking in bookings:
            if booking.get("created_at") and booking.get("payment_status") == "paid":
                try:
                    month = booking["created_at"][:7]  # YYYY-MM
                    monthly_data[month]["count"] += 1
                    monthly_data[month]["revenue"] += float(
                        booking.get("total_price", 0)
                    )
                except:
                    pass

        for month in sorted(monthly_data.keys(), reverse=True)[:12]:
            writer.writerow(
                [
                    month,
                    monthly_data[month]["count"],
                    f"₹{monthly_data[month]['revenue']:,.2f}",
                ]
            )

        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "text/csv"
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        return response

    except Exception as e:
        print(f"Error exporting revenue: {e}")
        import traceback

        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.errorhandler(404)
def page_not_found(e):
    return render_template("404.html", user=session.get("user_info", {})), 404


# ============================================================================
# BOOKING MODIFICATIONS (Self-Service)
# ============================================================================


@app.route("/booking/<booking_id>/modify", methods=["GET", "POST"])
@login_required
def modify_booking(booking_id):
    """Allow guests to modify their bookings"""
    booking = get_booking(booking_id)

    if not booking:
        flash("Booking not found", "danger")
        return redirect(url_for("my_bookings"))

    # Check if user owns this booking
    if booking.get("user_id") != session.get("user_id"):
        flash("Unauthorized access", "danger")
        return redirect(url_for("my_bookings"))

    if request.method == "POST":
        try:
            new_check_in = request.form.get("check_in")
            new_check_out = request.form.get("check_out")

            if new_check_in and new_check_out:
                # Calculate new price
                room = next(
                    (
                        r
                        for r in get_all_rooms()
                        if r.get("room_id") == booking.get("room_id")
                    ),
                    None,
                )
                if room:
                    base_price = float(room.get("price", 0))
                    new_price = calculate_dynamic_price(
                        base_price,
                        new_check_in,
                        new_check_out,
                        booking.get("room_id"),
                        booking.get("branch_id", ""),
                    )

                    # Update booking
                    booking["check_in"] = new_check_in
                    booking["check_out"] = new_check_out
                    nights = (
                        datetime.strptime(new_check_out, "%Y-%m-%d")
                        - datetime.strptime(new_check_in, "%Y-%m-%d")
                    ).days
                    booking["nights"] = nights
                    booking["total_price"] = new_price
                    booking["modified_at"] = datetime.now().isoformat()

                    add_booking(booking)  # This will update in DynamoDB

                    flash("Booking modified successfully!", "success")
                    return redirect(url_for("my_bookings"))

        except Exception as e:
            print(f"Error modifying booking: {e}")
            flash(f"Error modifying booking: {str(e)}", "danger")

    return render_template(
        "modify_booking.html", booking=booking, user=session.get("user_info", {})
    )


@app.route("/booking/<booking_id>/cancel", methods=["POST"])
@login_required
def cancel_booking_self_service(booking_id):
    """Allow guests to cancel their own bookings"""
    booking = get_booking(booking_id)

    if not booking:
        return jsonify({"success": False, "error": "Booking not found"}), 404

    # Check if user owns this booking
    if booking.get("user_id") != session.get("user_id"):
        return jsonify({"success": False, "error": "Unauthorized"}), 403

    try:
        # Calculate refund based on cancellation policy
        check_in = datetime.strptime(booking.get("check_in"), "%Y-%m-%d")
        days_until_checkin = (check_in - datetime.now()).days
        total_price = float(booking.get("total_price", 0))

        # Cancellation policy
        if days_until_checkin >= 7:
            refund_percent = 100  # Full refund
        elif days_until_checkin >= 3:
            refund_percent = 50  # 50% refund
        else:
            refund_percent = 0  # No refund

        refund_amount = total_price * (refund_percent / 100)
        cancellation_fee = total_price - refund_amount

        # Update booking status
        update_booking_status(booking_id, "cancelled")

        # Free up the room
        room_id = booking.get("room_id")
        if room_id:
            update_room_availability(room_id, "available")
            # Notify waitlist
            notify_waitlist(room_id)

        return jsonify(
            {
                "success": True,
                "message": "Booking cancelled successfully",
                "refund_amount": refund_amount,
                "cancellation_fee": cancellation_fee,
                "refund_percent": refund_percent,
            }
        )
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================================================
# WAITLIST SYSTEM
# ============================================================================


@app.route("/room/<room_id>/join-waitlist", methods=["POST"])
@login_required
def join_waitlist(room_id):
    """Join waitlist for unavailable room"""
    try:
        room = next((r for r in get_all_rooms() if r.get("room_id") == room_id), None)
        if not room:
            return jsonify({"success": False, "error": "Room not found"}), 404

        check_in = request.form.get("check_in")
        check_out = request.form.get("check_out")

        waitlist_data = {
            "waitlist_id": str(uuid.uuid4()),
            "user_id": session["user_id"],
            "room_id": room_id,
            "room_name": room.get("name", "Unknown Room"),
            "branch_id": room.get("branch_id"),
            "check_in": check_in,
            "check_out": check_out,
            "status": "active",
            "created_at": datetime.now().isoformat(),
        }

        add_to_waitlist(waitlist_data)

        flash(
            "You have been added to the waitlist! We will notify you when this room becomes available.",
            "success",
        )
        return jsonify({"success": True, "message": "Added to waitlist"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/my-waitlist")
@login_required
def my_waitlist():
    """View user's waitlist entries"""
    try:
        user_id = session.get("user_id")
        waitlist = get_user_waitlist(user_id)
        return render_template(
            "waitlist.html", waitlist=waitlist, user=session.get("user_info", {})
        )
    except Exception as e:
        print(f"Error loading waitlist: {e}")
        flash(f"Error loading waitlist: {str(e)}", "danger")
        return render_template(
            "waitlist.html", waitlist=[], user=session.get("user_info", {})
        )


@app.route("/waitlist/<waitlist_id>/remove", methods=["POST"])
@login_required
def remove_from_waitlist(waitlist_id):
    """Remove from waitlist"""
    try:
        update_waitlist_status(waitlist_id, "cancelled")
        return jsonify({"success": True, "message": "Removed from waitlist"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ============================================================================
# SERVICE BOOKINGS (Spa, Restaurant, Transportation, etc.)
# ============================================================================


@app.route("/services")
@login_required
def services_list():
    """View available services"""
    try:
        user_info = session.get("user_info", {})
        branch_id = request.args.get("branch_id")

        services = get_services(branch_id=branch_id)
        branches = get_all_branches()

        return render_template(
            "services.html", services=services, branches=branches, user=user_info
        )
    except Exception as e:
        print(f"Error loading services: {e}")
        flash(f"Error loading services: {str(e)}", "danger")
        return render_template(
            "services.html", services=[], branches=[], user=session.get("user_info", {})
        )


@app.route("/service/<service_id>/book", methods=["POST"])
@login_required
def book_service_route(service_id):
    """Book a service"""
    try:
        services = get_services()
        service = next((s for s in services if s.get("service_id") == service_id), None)

        if not service:
            return jsonify({"success": False, "error": "Service not found"}), 404

        booking_date = request.form.get("booking_date")
        booking_time = request.form.get("booking_time")
        booking_id = request.form.get("booking_id")  # Optional: link to room booking
        notes = request.form.get("notes", "")

        service_booking_data = {
            "service_booking_id": str(uuid.uuid4()),
            "user_id": session["user_id"],
            "service_id": service_id,
            "service_name": service.get("name"),
            "service_type": service.get("service_type"),
            "booking_id": booking_id,
            "booking_date": booking_date,
            "booking_time": booking_time,
            "price": service.get("price"),
            "status": "confirmed",
            "notes": notes,
            "created_at": datetime.now().isoformat(),
        }

        book_service(service_booking_data)

        flash(f'{service.get("name")} booked successfully!', "success")
        return jsonify({"success": True, "message": "Service booked successfully"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/my-service-bookings")
@login_required
def my_service_bookings():
    """View user's service bookings"""
    try:
        user_id = session.get("user_id")
        service_bookings = get_user_service_bookings(user_id)
        return render_template(
            "service_bookings.html",
            service_bookings=service_bookings,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading service bookings: {e}")
        flash(f"Error loading service bookings: {str(e)}", "danger")
        return render_template(
            "service_bookings.html",
            service_bookings=[],
            user=session.get("user_info", {}),
        )


# ============================================================================
# GUEST PREFERENCES
# ============================================================================


@app.route("/preferences", methods=["GET", "POST"])
@login_required
def guest_preferences():
    """Manage guest preferences"""
    user_id = session.get("user_id")

    if request.method == "POST":
        try:
            preferences = {
                "room_type_preference": request.form.get("room_type_preference"),
                "floor_preference": request.form.get("floor_preference"),
                "bed_type": request.form.get("bed_type"),
                "view_preference": request.form.get("view_preference"),
                "dietary_restrictions": request.form.get("dietary_restrictions"),
                "special_requests": request.form.get("special_requests"),
                "accessibility_needs": request.form.get("accessibility_needs"),
                "smoking_preference": request.form.get("smoking_preference"),
                "pillow_preference": request.form.get("pillow_preference"),
                "temperature_preference": request.form.get("temperature_preference"),
            }

            save_guest_preferences(user_id, preferences)
            flash("Preferences saved successfully!", "success")
            return redirect(url_for("guest_preferences"))
        except Exception as e:
            print(f"Error saving preferences: {e}")
            flash(f"Error saving preferences: {str(e)}", "danger")

    # GET request
    try:
        preferences = get_guest_preferences(user_id)
        return render_template(
            "preferences.html",
            preferences=preferences,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading preferences: {e}")
        flash(f"Error loading preferences: {str(e)}", "danger")
        return render_template(
            "preferences.html", preferences={}, user=session.get("user_info", {})
        )


# ============================================================================
# ROOM COMPARISON
# ============================================================================


@app.route("/compare-rooms")
def compare_rooms():
    """Compare multiple rooms side by side"""
    try:
        room_ids = request.args.getlist("room_ids")

        if not room_ids or len(room_ids) < 2:
            flash("Please select at least 2 rooms to compare", "warning")
            return redirect(url_for("rooms"))

        if len(room_ids) > 4:
            flash("You can compare up to 4 rooms at a time", "warning")
            room_ids = room_ids[:4]

        all_rooms = get_all_rooms()
        rooms_to_compare = [r for r in all_rooms if r.get("room_id") in room_ids]

        return render_template(
            "compare_rooms.html",
            rooms=rooms_to_compare,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error comparing rooms: {e}")
        flash(f"Error comparing rooms: {str(e)}", "danger")
        return redirect(url_for("rooms"))


# ============================================================================
# NOTIFICATIONS
# ============================================================================


@app.route("/notifications")
@login_required
def user_notifications():
    """View user notifications"""
    try:
        user_id = session.get("user_id")
        notifications = get_user_notifications(user_id)
        return render_template(
            "notifications.html",
            notifications=notifications,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading notifications: {e}")
        flash(f"Error loading notifications: {str(e)}", "danger")
        return render_template(
            "notifications.html", notifications=[], user=session.get("user_info", {})
        )


# ============================================================================
# PRICING MANAGEMENT (Admin Only)
# ============================================================================


@app.route("/admin/pricing-rules")
@login_required
@role_required("admin")
def pricing_rules_management():
    """Manage dynamic pricing rules"""
    try:
        rules = get_pricing_rules()
        branches = get_all_branches()
        return render_template(
            "pricing_rules.html",
            rules=rules,
            branches=branches,
            user=session.get("user_info", {}),
        )
    except Exception as e:
        print(f"Error loading pricing rules: {e}")
        flash(f"Error loading pricing rules: {str(e)}", "danger")
        return render_template(
            "pricing_rules.html",
            rules=[],
            branches=[],
            user=session.get("user_info", {}),
        )


@app.errorhandler(500)
def internal_server_error(e):
    return render_template("500.html", user=session.get("user_info", {})), 500


@app.route("/404")
def error_404():
    return render_template("404.html", user=session.get("user_info", {}))


@app.route("/500")
def error_500():
    return render_template("500.html", user=session.get("user_info", {}))


# =============================================================================
#  AI / ML FEATURE ROUTES
# =============================================================================


# ── 1. AI Room Recommendations (Page + API) ───────────────────────────────────
@app.route("/ai-recommendations")
@login_required
def ai_recommendations_page():
    """Full AI recommendations page for guests."""
    try:
        user_id = session.get("user_id", "")
        user = get_user(user_id) or {}
        all_rooms = get_all_rooms() or []
        all_bookings = get_all_bookings() or []
        all_users = get_all_users() or []

        my_bookings = [b for b in all_bookings if b.get("user_id") == user_id]
        loyalty_tier = user.get("loyalty_tier", "Silver")

        recommendations = []
        if AI_RECOMMENDER_OK:
            recommendations = get_recommendations(
                user_id=user_id,
                all_rooms=all_rooms,
                past_bookings=my_bookings,
                all_bookings=all_bookings,
                all_users=all_users,
                loyalty_tier=loyalty_tier,
                top_n=6,
            )

        return render_template(
            "ai_recommendations.html",
            recommendations=recommendations,
            loyalty_tier=loyalty_tier,
            user=user,
            model=REC_META if AI_RECOMMENDER_OK else {},
        )
    except Exception as e:
        print(f"AI recommendations error: {e}")
        return render_template(
            "ai_recommendations.html",
            recommendations=[],
            loyalty_tier="Silver",
            user=session.get("user_info", {}),
            model={},
        )


@app.route("/api/recommendations")
@login_required
def api_recommendations():
    """JSON API for AJAX recommendation refresh."""
    try:
        user_id = session.get("user_id", "")
        user = get_user(user_id) or {}
        all_rooms = get_all_rooms() or []
        all_bookings = get_all_bookings() or []
        all_users = get_all_users() or []

        my_bookings = [b for b in all_bookings if b.get("user_id") == user_id]
        loyalty_tier = user.get("loyalty_tier", "Silver")
        top_n = int(request.args.get("top_n", 6))
        nights = int(request.args.get("nights", 2))

        if not AI_RECOMMENDER_OK:
            return jsonify(
                {"success": False, "error": "Recommender module not available"}
            )

        recs = get_recommendations(
            user_id=user_id,
            all_rooms=all_rooms,
            past_bookings=my_bookings,
            all_bookings=all_bookings,
            all_users=all_users,
            loyalty_tier=loyalty_tier,
            top_n=top_n,
            nights=nights,
        )

        # Flatten room into top-level keys for JS
        flat = []
        for r in recs:
            item = dict(r["room"])
            item["score"] = r["score"]
            item["confidence"] = r["confidence"]
            item["reasons"] = r["reasons"]
            item["match_label"] = r["match_label"]
            item["component_scores"] = r["component_scores"]
            flat.append(item)

        return jsonify({"success": True, "recommendations": flat, "count": len(flat)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── 2. Demand Forecasting API ─────────────────────────────────────────────────
@app.route("/api/ai/demand-forecast")
@login_required
def api_demand_forecast():
    """30-day demand / occupancy forecast (managers & above)."""
    try:
        role = session.get("role", "guest")
        if role not in ("manager", "admin", "super_admin"):
            return jsonify({"success": False, "error": "Access denied"}), 403

        if not AI_FORECAST_OK:
            return jsonify({"success": False, "error": "Forecast module not available"})

        all_bookings = get_all_bookings() or []
        all_rooms = get_all_rooms() or []
        result = forecast_demand(all_bookings, total_rooms=len(all_rooms) or 100)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── 3. Dynamic Pricing API ────────────────────────────────────────────────────
@app.route("/api/ai/dynamic-price/<room_id>")
def api_dynamic_price(room_id):
    """Get dynamic price for a single room."""
    try:
        if not AI_PRICING_OK:
            return jsonify({"success": False, "error": "Pricing module not available"})

        all_rooms = get_all_rooms() or []
        all_bookings = get_all_bookings() or []
        room = next((r for r in all_rooms if r.get("room_id") == room_id), None)
        if not room:
            # Use a demo room for pricing simulation (useful when DB is empty)
            room = {
                "room_id": room_id,
                "room_type": "double",
                "price": 5000,
                "name": "Demo Room",
            }

        occupied = sum(
            1
            for b in all_bookings
            if b.get("booking_status", b.get("status")) == "confirmed"
        )
        occ_rate = occupied / max(len(all_rooms), 1)
        days = int(request.args.get("days_until", 7))

        pricing = compute_dynamic_price(room, occ_rate, days_until_checkin=days)
        return jsonify({"success": True, **pricing})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── 4. Sentiment Analysis API ─────────────────────────────────────────────────
@app.route("/api/ai/sentiment")
@login_required
def api_sentiment():
    """Batch sentiment analysis on all reviews."""
    try:
        role = session.get("role", "guest")
        if role not in ("manager", "admin", "super_admin"):
            return jsonify({"success": False, "error": "Access denied"}), 403

        if not AI_SENTIMENT_OK:
            return jsonify(
                {"success": False, "error": "Sentiment module not available"}
            )

        all_reviews = get_all_reviews() or []
        result = analyze_all_reviews(all_reviews)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── 5. Fraud Detection API ────────────────────────────────────────────────────
@app.route("/api/ai/fraud-scan")
@login_required
def api_fraud_scan():
    """Scan all bookings for fraud signals."""
    try:
        role = session.get("role", "guest")
        if role not in ("admin", "super_admin"):
            return jsonify({"success": False, "error": "Access denied"}), 403

        if not AI_FRAUD_OK:
            return jsonify(
                {"success": False, "error": "Fraud detection module not available"}
            )

        all_bookings = get_all_bookings() or []
        all_rooms = get_all_rooms() or []
        all_users = get_all_users() or []
        result = scan_all_bookings(all_bookings, all_rooms, all_users)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route("/api/ai/fraud-score/<booking_id>")
@login_required
def api_fraud_score_single(booking_id):
    """Get fraud score for a single booking."""
    try:
        role = session.get("role", "guest")
        if role not in ("admin", "super_admin"):
            return jsonify({"success": False, "error": "Access denied"}), 403

        if not AI_FRAUD_OK:
            return jsonify(
                {"success": False, "error": "Fraud detection module not available"}
            )

        all_bookings = get_all_bookings() or []
        all_rooms = get_all_rooms() or []
        all_users = get_all_users() or []

        booking = next(
            (b for b in all_bookings if b.get("booking_id") == booking_id), None
        )
        if not booking:
            return jsonify({"success": False, "error": "Booking not found"}), 404

        room = next(
            (r for r in all_rooms if r.get("room_id") == booking.get("room_id")), {}
        )
        user = next(
            (u for u in all_users if u.get("user_id") == booking.get("user_id")), {}
        )
        result = fraud_score_booking(booking, room, user, all_bookings)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── 6. Cancellation Prediction API ───────────────────────────────────────────
@app.route("/api/ai/cancellation-risk")
@login_required
def api_cancellation_risk():
    """Predict cancellation risk for all active bookings."""
    try:
        role = session.get("role", "guest")
        if role not in ("manager", "admin", "super_admin"):
            return jsonify({"success": False, "error": "Access denied"}), 403

        if not AI_CANCEL_OK:
            return jsonify(
                {"success": False, "error": "Cancellation module not available"}
            )

        all_bookings = get_all_bookings() or []
        all_users = get_all_users() or []
        result = predict_all_cancellations(all_bookings, all_users)
        return jsonify({"success": True, **result})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ── AI Reservation Agent ────────────────────────────────────────────────────────
@app.route("/chatbot")
def ai_agent_chatbot():
    """Renders the embedded Streamlit chatbot for hotel search and booking."""
    return render_template("ai_agent.html", role=session.get("role", "guest"))


# ── AI Dashboard Overview Page ────────────────────────────────────────────────
@app.route("/ai-dashboard")
@login_required
def ai_dashboard():
    """Central AI/ML overview page — accessible to managers and above."""
    try:
        role = session.get("role", "guest")
        if role == "guest":
            return redirect(url_for("ai_recommendations_page"))

        user = get_user(session.get("user_id", "")) or {}
        all_bookings = get_all_bookings() or []
        all_rooms = get_all_rooms() or []
        all_users = get_all_users() or []
        all_reviews = get_all_reviews() or []

        # Quick stats for AI dashboard header
        ai_stats = {
            "total_rooms": len(all_rooms),
            "total_bookings": len(all_bookings),
            "total_reviews": len(all_reviews),
            "total_users": len(all_users),
            "modules_active": sum(
                [
                    AI_RECOMMENDER_OK,
                    AI_FORECAST_OK,
                    AI_PRICING_OK,
                    AI_SENTIMENT_OK,
                    AI_FRAUD_OK,
                    AI_CANCEL_OK,
                ]
            ),
        }

        # Sentiment quick summary
        sentiment_summary = analyze_all_reviews(all_reviews) if AI_SENTIMENT_OK else {}

        # Cancellation risk summary
        cancel_summary = (
            predict_all_cancellations(all_bookings, all_users) if AI_CANCEL_OK else {}
        )

        # Fraud scan summary
        fraud_summary = (
            scan_all_bookings(all_bookings, all_rooms, all_users) if AI_FRAUD_OK else {}
        )

        return render_template(
            "ai_dashboard.html",
            user=user,
            role=role,
            ai_stats=ai_stats,
            sentiment_summary=sentiment_summary,
            cancel_summary=cancel_summary,
            fraud_summary=fraud_summary,
            modules={
                "recommender": AI_RECOMMENDER_OK,
                "forecast": AI_FORECAST_OK,
                "pricing": AI_PRICING_OK,
                "sentiment": AI_SENTIMENT_OK,
                "fraud": AI_FRAUD_OK,
                "cancellation": AI_CANCEL_OK,
            },
        )
    except Exception as e:
        print(f"AI Dashboard error: {e}")
        import traceback

        traceback.print_exc()
        return redirect(url_for("guest_dashboard"))


if __name__ == "__main__":
    # Initialize database on startup
    init_db()

    # Print startup info
    print("\n" + "=" * 50)
    print("Blissful Abodes Hotel Booking System")
    print("=" * 50)
    print(f"Server running on: http://localhost:5000")
    print("\nTest Accounts:")
    print("  Guest: guest@example.com / password123")
    print("  Staff: staff@example.com / password123")
    print("  Admin: admin@example.com / password123")
    print("\nPress Ctrl+C to stop the server")
    print("=" * 50 + "\n")

    # Run the app
    app.run(host="0.0.0.0", port=5000, debug=True)
